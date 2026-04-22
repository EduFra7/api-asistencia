# ==============================================================================
# 1. IMPORTACIÓN DE LIBRERÍAS (Las herramientas que usamos)
# ==============================================================================

# -- Framework Web (FastAPI) --
from fastapi import FastAPI, Request, HTTPException, Depends, BackgroundTasks
from fastapi.responses import PlainTextResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer

# -- Manejo de Tiempo y Fechas (Nativas y Externas) --
from datetime import datetime, date, timedelta, time
from dateutil.relativedelta import relativedelta  # Instalado vía pip

# -- Seguridad y Base de Datos --
import psycopg2          # Librería para hablar con PostgreSQL (Supabase)
import psycopg2.extras   # Herramientas extra para PostgreSQL (como diccionarios)
import bcrypt            # Librería para encriptar contraseñas
import jwt               # Librería para generar "Tokens" de sesión

# -- Utilidades del Sistema --
from dotenv import load_dotenv
import os                # Librería para leer variables del sistema operativo
import re                # Librería para buscar y limpiar textos (Expresiones regulares)
import json              # Asegúrate de que esto esté al inicio de tu main.py

# -- Librerías para feriados --
import holidays
from pydantic import BaseModel
from typing import Optional

# -- Librerias para reportes PDF y EXCEL --
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A5, portrait
from reportlab.pdfgen import canvas
from reportlab.lib import colors

# ==============================================================================
# 2. CONFIGURACIÓN INICIAL DE LA APLICACIÓN
# ==============================================================================
# load_dotenv() busca un archivo llamado '.env' y carga sus variables secretas.
load_dotenv() 

# app = FastAPI() es la línea más importante. Crea el servidor web.
app = FastAPI()

# ── CORS (Cross-Origin Resource Sharing) ──
# El CORS es un guardia de seguridad del navegador. Por defecto, un navegador no 
# permite que una web (index.html) pida datos a una API que está en otra dirección.
# Al poner allow_origins=["*"], le decimos a la API: "Acepta peticiones de cualquier página web".
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Obtenemos la llave maestra para firmar tokens. Si no existe en el .env, usa una por defecto.
SECRET_KEY = os.getenv("SECRET_KEY", "clave_secreta_cambiar_en_produccion")


# ==============================================================================
# 3. FUNCIONES CORE (Funciones que se reutilizan en todo el código)
# ==============================================================================

# ── CONEXIÓN DINÁMICA A LA BASE DE DATOS ──
def conectar_bd(schema_name="public"):
    """
    Esta función abre el puente entre Python y Supabase.
    El parámetro 'schema_name' es crucial para el multitenant: le dice a la base de datos
    a qué "carpeta" privada debe entrar antes de buscar tablas.
    """
    url = os.getenv("DATABASE_URL")
    if not url:
        raise Exception("DATABASE_URL no encontrada")
    
    # search_path es un comando de PostgreSQL que le indica qué esquema usar.
    return psycopg2.connect(url, options=f"-c search_path={schema_name}")


# ── VERIFICACIÓN DE IDENTIDAD (MIDDLEWARE DE SEGURIDAD) ──
def verificar_token(request: Request):
    """
    Cada vez que un usuario intenta entrar a una ruta protegida (ej. crear empresa),
    esta función intercepta la petición, busca el 'token' en la cabecera, 
    y lo descifra para ver quién es.
    """
    # Extraemos el token que el frontend manda oculto en las cabeceras.
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        raise HTTPException(status_code=401, detail="Token requerido")
    
    try:
        # jwt.decode() intenta abrir el token con nuestra llave maestra.
        # Si alguien alteró el token, la firma no coincidirá y fallará.
        datos = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return datos # Retorna el diccionario con los datos del usuario (id, rol, schema, etc.)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado. Inicie sesión de nuevo.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido o manipulado.")

# ── HORA OFICIAL DEL SERVIDOR (Para sincronizar Frontends) ──
@app.get("/hora-servidor")
def obtener_hora_servidor():
    # Devuelve la hora exacta de Railway (con la zona horaria America/La_Paz que configuraste)
    return {"hora_oficial": datetime.now().isoformat()}

# ==============================================================================
# 4. RUTAS O ENDPOINTS (Las "puertas" de tu servidor)
# ==============================================================================

# ── SISTEMA DE LOGIN ──
# @app.post significa que esta ruta espera recibir datos (en este caso, email y password)
@app.post("/auth/login")
async def login(request: Request):
    # 'async' permite que el servidor no se congele mientras espera datos.
    try:
        data     = await request.json() # Leemos los datos que mandó el navegador
        email    = data.get("email", "").strip().lower()
        password = data.get("password", "")

        # 1. Buscamos al usuario siempre en la tabla global ('public')
        conn = conectar_bd("public")
        # RealDictCursor hace que los resultados salgan como diccionarios {"nombre": "Juan"}
        # en lugar de simples tuplas ("Juan",)
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # Hacemos un JOIN para traer tanto los datos del usuario como el esquema de su empresa
        cur.execute("""
            SELECT u.*, e.nombre as empresa_nombre, e.schema_name
            FROM usuarios u
            JOIN empresas e ON e.id = u.empresa_id
            WHERE u.email = %s AND u.activo = TRUE
        """, (email,))
        usuario = cur.fetchone() # Trae el primer resultado encontrado
        cur.close()
        conn.close()

        # 2. Validaciones de seguridad
        if not usuario:
            raise HTTPException(status_code=401, detail="Credenciales incorrectas")

        # bcrypt.checkpw() compara la contraseña que escribió el usuario con el hash raro de la BD.
        if not bcrypt.checkpw(password.encode(), usuario["password_hash"].encode()):
            raise HTTPException(status_code=401, detail="Credenciales incorrectas")

        # 3. Creación del "Pasaporte" (Token JWT)
        # Este token viaja al navegador y es la única prueba de que el usuario ya inició sesión.
        token = jwt.encode({
            "id":             usuario["id"],
            "email":          usuario["email"],
            "rol":            usuario["rol"],
            "empresa_id":     usuario["empresa_id"],
            "schema_name":    usuario["schema_name"], # Esto es vital para saber a qué esquema enviarlo luego
            "empresa_nombre": usuario["empresa_nombre"],
            "exp":            datetime.utcnow() + timedelta(hours=8) # El token caduca en 8 horas
        }, SECRET_KEY, algorithm="HS256")

        # Devolvemos el token y los datos básicos para que el frontend arme la interfaz
        return {
            "token":          token,
            "nombre":         usuario["nombre"],
            "rol":            usuario["rol"],
            "empresa_nombre": usuario["empresa_nombre"],
            "schema_name":    usuario["schema_name"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── RUTA DE PRUEBA DE IDENTIDAD ──
# Observa el 'Depends(verificar_token)'. Esto fuerza a que la función verificar_token()
# se ejecute ANTES de entrar a esta ruta. Si el token falla, nunca llega aquí.
@app.get("/auth/me")
def mi_perfil(usuario = Depends(verificar_token)):
    return usuario # Devuelve lo que la función verificar_token descifró


# ── MOTOR DE APROVISIONAMIENTO (CREAR EMPRESAS NUEVAS) ──
@app.post("/empresas")
async def crear_empresa(request: Request, usuario = Depends(verificar_token)):
    # 1. Filtro de seguridad: ¿Es el dueño del sistema?
    if usuario["rol"] != "superadmin":
        raise HTTPException(status_code=403, detail="Sin permisos. Solo SuperAdmin.")
    
    try:
        data = await request.json()
        nombre         = data.get("nombre")
        admin_nombre   = data.get("admin_nombre")
        admin_email    = data.get("admin_email")
        admin_password = data.get("admin_password")

        # 2. Preparar un nombre seguro para el esquema en PostgreSQL (sin espacios ni símbolos)
        schema_limpio = re.sub(r'[^a-z0-9_]', '', nombre.lower().replace(' ', '_'))
        schema_name   = f"empresa_{schema_limpio}" 

        # Encriptamos la contraseña del cliente antes de guardarla
        password_hash = bcrypt.hashpw(admin_password.encode(), bcrypt.gensalt()).decode()

        conn = conectar_bd("public")
        # autocommit=True es necesario porque PostgreSQL no deja crear esquemas dentro de una "transacción" normal.
        conn.autocommit = True 
        cur  = conn.cursor()

        # 3. Guardar en el directorio maestro de empresas
        cur.execute(
            "INSERT INTO empresas (nombre, schema_name) VALUES (%s, %s) RETURNING id",
            (nombre, schema_name)
        )
        empresa_id = cur.fetchone()[0] # Obtenemos el ID que se le asignó

        # 4. Crear al Administrador principal de este nuevo cliente
        cur.execute("""
            INSERT INTO usuarios (empresa_id, nombre, email, password_hash, rol)
            VALUES (%s, %s, %s, %s, 'admin')
        """, (empresa_id, admin_nombre, admin_email, password_hash))

        # 5. LA MAGIA: Crear el esquema y sus tablas privadas (VERSIÓN 2.0)
        from psycopg2 import sql
        
        # ATENCIÓN: El orden de creación importa. Primero los catálogos (sin dependencias), 
        # luego los empleados, y al final los eventos.
        script_sql = sql.SQL("""
            CREATE SCHEMA {schema};

            -- ==========================================
            -- 1. TABLAS DE CATÁLOGOS (Datos Fijos)
            -- ==========================================
                             
            -- ⚡ NUEVA TABLA DE FERIADOS
            CREATE TABLE {schema}.feriados (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                descripcion VARCHAR(150) NOT NULL,
                tipo VARCHAR(50) DEFAULT 'nacional',
                recurrente BOOLEAN DEFAULT FALSE,
                eliminado BOOLEAN DEFAULT FALSE
            );
            
            INSERT INTO {schema}.feriados (fecha, descripcion, tipo, recurrente) VALUES
            ('2026-01-01', 'Año Nuevo', 'nacional', TRUE),
            ('2026-01-22', 'Día del Estado Plurinacional', 'nacional', TRUE),
            ('2026-02-16', 'Lunes de Carnaval', 'nacional', FALSE),
            ('2026-02-17', 'Martes de Carnaval', 'nacional', FALSE),
            ('2026-04-03', 'Viernes Santo', 'nacional', FALSE),
            ('2026-05-01', 'Día del Trabajo', 'nacional', TRUE),
            ('2026-06-04', 'Corpus Christi', 'nacional', FALSE),
            ('2026-06-21', 'Año Nuevo Aymara', 'nacional', TRUE),
            ('2026-08-06', 'Día de la Independencia', 'nacional', TRUE),
            ('2026-11-02', 'Todos Santos', 'nacional', TRUE),
            ('2026-12-25', 'Navidad', 'nacional', TRUE);
                             
            CREATE TABLE {schema}.sucursales (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(100) NOT NULL,
                ciudad VARCHAR(50) DEFAULT 'Nacional', -- ⚡ NUEVA COLUMNA
                direccion TEXT,
                telefono VARCHAR(50) DEFAULT '',
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                eliminado BOOLEAN DEFAULT FALSE
            );

            -- ¡NUEVA TABLA DE SECCIONES!
            CREATE TABLE {schema}.secciones (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(100) NOT NULL,
                descripcion TEXT,
                estado BOOLEAN DEFAULT TRUE,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                eliminado BOOLEAN DEFAULT FALSE
            );
             
            CREATE TABLE {schema}.turnos (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(100) NOT NULL,
                hora_ingreso TIME NOT NULL,
                hora_salida TIME NOT NULL,
                dias JSONB NOT NULL,
                almuerzo BOOLEAN DEFAULT FALSE,
                hora_inicio_almuerzo TIME,       -- ⚡ NUEVO
                hora_fin_almuerzo TIME,          -- ⚡ NUEVO
                almuerzo_min INTEGER DEFAULT 0,
                tolerancia BOOLEAN DEFAULT FALSE,
                tolerancia_min INTEGER DEFAULT 0,
                tolerancia_mensual_min INTEGER DEFAULT 0,
                descuento BOOLEAN DEFAULT TRUE,
                horas_extras BOOLEAN DEFAULT FALSE,
                medio_tiempo_fines BOOLEAN DEFAULT FALSE,
                eliminado BOOLEAN DEFAULT FALSE
            )

            -- ==========================================
            -- 2. TABLA PRINCIPAL DE EMPLEADOS (Planilla)
            -- ==========================================
            CREATE TABLE {schema}.empleados (
                id SERIAL PRIMARY KEY,
                bio_id INT UNIQUE,                     -- ID Numérico del Lector Biométrico
                foto_perfil TEXT,                      -- ⚡ NUEVO: CAMPO PARA LA FOTO BASE64
                nombres VARCHAR(100) NOT NULL,
                apellidos VARCHAR(100) NOT NULL,
                ci VARCHAR(30) UNIQUE NOT NULL,        -- Carnet de Identidad (único)
                sexo VARCHAR(20),
                celular VARCHAR(20),
                correo VARCHAR(100),
                direccion TEXT,
                fecha_ingreso DATE,
                fecha_antiguedad DATE,                 -- Para cálculo de vacaciones/bonos
                cargo VARCHAR(100),                    -- Texto libre (apoyado por el datalist del HTML)
                sucursal_id INT REFERENCES {schema}.sucursales(id), -- Conexión Física
                seccion_id INT REFERENCES {schema}.secciones(id),   -- ¡NUEVA! Conexión Lógica
                tipo_contrato VARCHAR(50),
                turno_id INT REFERENCES {schema}.turnos(id),      -- Conexión a su horario
                salario_base NUMERIC(10, 2) DEFAULT 0.00,           -- Permite decimales (Ej. 2500.50)
                bono NUMERIC(10, 2) DEFAULT 0.00,
                
                -- Control de Estado (Soft Delete y Bajas)
                activo BOOLEAN DEFAULT TRUE,           -- El interruptor principal
                fecha_retiro DATE,                     -- Se llena solo si activo pasa a FALSE
                motivo_retiro TEXT,
                eliminado BOOLEAN DEFAULT FALSE,       -- Para el botón "Eliminar" (Borrado Lógico)
                
                huella_template TEXT,                  -- Aquí guardaremos el string de la huella dactilar
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                actualizado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                historial_movimientos TEXT DEFAULT ''   -- Caja negra de registro de movimiento
            );

            -- ==========================================
            -- 3. TABLAS DE HARDWARE Y ASISTENCIA
            -- ==========================================
            CREATE TABLE {schema}.eventos_brutos (
                id SERIAL PRIMARY KEY,
                device_no VARCHAR(50),                 -- Número de serie del reloj ZKTeco
                item VARCHAR(50),                      -- El bio_id del empleado que marcó
                verify_mode VARCHAR(20),               -- Huella, Tarjeta, Rostro o Clave
                action VARCHAR(20),                    -- Entrada, Salida, etc.
                fecha_hora TIMESTAMP,                  -- El momento exacto del marcaje
                raw_data JSONB,                        -- El paquete de datos original por si hay que depurar
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
                             
            -- ⚡ NUEVA TABLA MAESTRA DE CÁLCULO DIARIO ⚡
            CREATE TABLE {schema}.asistencia_diaria (
                id SERIAL PRIMARY KEY, -- Usamos SERIAL para mantener el estándar de tu BD
                empleado_id INT REFERENCES {schema}.empleados(id) NOT NULL,
                fecha DATE NOT NULL,
                turno_id INT REFERENCES {schema}.turnos(id), 
                
                -- Marcajes limpios
                hora_entrada TIME,
                hora_inicio_almuerzo TIME,
                hora_fin_almuerzo TIME,
                hora_salida TIME,
                
                -- Desglose de Cálculos
                minutos_retraso_entrada INT DEFAULT 0,
                minutos_exceso_almuerzo INT DEFAULT 0,
                minutos_salida_temprano INT DEFAULT 0,
                horas_trabajadas NUMERIC(5,2) DEFAULT 0.00,
                horas_extras NUMERIC(5,2) DEFAULT 0.00,
                
                -- Veredicto y Finanzas
                estado VARCHAR(20) NOT NULL DEFAULT 'Incompleto', 
                deuda_generada_bs NUMERIC(10, 2) DEFAULT 0.00, 
                
                -- Auditoría
                modificado_manualmente BOOLEAN DEFAULT FALSE,
                observaciones TEXT,
                actualizado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                
                UNIQUE(empleado_id, fecha) -- Evita duplicados del mismo día
            );

            CREATE TABLE {schema}.asistencia (
                id SERIAL PRIMARY KEY,
                empleado_id INT REFERENCES {schema}.empleados(id),
                fecha DATE,
                hora_marcaje TIMESTAMP,
                tipo VARCHAR(20),                      -- Ej: 'Entrada_Normal', 'Salida_Almuerzo'
                estado VARCHAR(20),                    -- Ej: 'A_Tiempo', 'Retraso', 'Falta'
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
                             
            -- ⚡ ¡AQUÍ AGREGAMOS LA TABLA QUE FALTABA PARA LAS VACACIONES! ⚡
            CREATE TABLE {schema}.ausencias (
                id SERIAL PRIMARY KEY,
                empleado_id INT REFERENCES {schema}.empleados(id) NOT NULL,
                tipo VARCHAR(20) NOT NULL,
                fecha_inicio DATE NOT NULL,
                fecha_fin DATE NOT NULL,
                hora_inicio TIME,
                hora_fin TIME,
                horas_totales NUMERIC(5,2) DEFAULT 0.00,
                dias_descontados NUMERIC(5,2) DEFAULT 0.00,
                motivo TEXT,
                requiere_reposicion BOOLEAN DEFAULT FALSE,
                estado VARCHAR(20) DEFAULT 'aprobado',
                eliminado BOOLEAN DEFAULT FALSE,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """).format(schema=sql.Identifier(schema_name))

        cur.execute(script_sql) # Ejecutamos todo el bloque para crear la empresa

        cur.close()
        conn.close()

        return {
            "mensaje": "Empresa aprovisionada correctamente", 
            "empresa_id": empresa_id, 
            "schema_name": schema_name
        }

    except psycopg2.errors.DuplicateSchema:
        raise HTTPException(status_code=400, detail="Ya existe una empresa con un nombre similar")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


# ── LISTAR EMPRESAS ──
@app.get("/empresas")
def ver_empresas(usuario = Depends(verificar_token)):
    if usuario["rol"] != "superadmin":
        raise HTTPException(status_code=403, detail="Sin permisos")
    
    conn = conectar_bd("public")
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM empresas ORDER BY creado_en DESC")
    empresas = cur.fetchall()
    cur.close()
    conn.close()
    return list(empresas)


# ── SETUP INICIAL (CREAR TU CUENTA) ──
@app.get("/setup/superadmin")
def crear_superadmin():
    """Esta ruta se usa solo una vez para crear al dueño del sistema."""
    try:
        password_hash = bcrypt.hashpw("admin123".encode(), bcrypt.gensalt()).decode()
        conn = conectar_bd("public")
        cur  = conn.cursor()
        
        # ON CONFLICT DO NOTHING evita que el código se rompa si "Sistema" ya existe
        cur.execute("INSERT INTO empresas (id, nombre, schema_name) VALUES (1, 'Sistema', 'public') ON CONFLICT (id) DO NOTHING")
        
        cur.execute("""
            INSERT INTO usuarios (empresa_id, nombre, email, password_hash, rol)
            VALUES (1, 'Super Admin', 'admin@sistema.com', %s, 'superadmin')
            ON CONFLICT (email) DO UPDATE SET password_hash = %s
        """, (password_hash, password_hash))
        
        conn.commit() # Guardar cambios en la base de datos
        cur.close()
        conn.close()
        return {"mensaje": "Superadmin creado. Email: admin@sistema.com / Pass: admin123"}
    except Exception as e:
        return {"error": str(e)}


# ==============================================================================
# X. CEREBRO DE CECÁLCULO - EVENT - DRIVEN
# ==============================================================================

# ⚡ FUNCIÓN MAESTRA: EL CEREBRO EVENT-DRIVEN
def procesar_asistencia_dia(schema: str, empleado_id: int, fecha: date):
    """
    Esta función es el único lugar donde se calcula la asistencia.
    Se dispara ante CUALQUIER evento (Huella, Permiso, Edición).
    """
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        # 1. Obtener datos básicos (Empleado, Turno y Salario)
        # ⚡ OPTIMIZACIÓN: Traemos el bio_id aquí directamente
        cur.execute(f"""
            SELECT e.id, e.bio_id, e.salario_base, e.turno_id, t.* FROM {schema}.empleados e 
            JOIN {schema}.turnos t ON e.turno_id = t.id 
            WHERE e.id = %s
        """, (empleado_id,))
        emp_turno = cur.fetchone()
        if not emp_turno: return False
        
        bio_id_str = str(emp_turno['bio_id']) if emp_turno.get('bio_id') else "S/N"

        # 2. 🌙 VENTANA DINÁMICA: Detección de Turno Nocturno
        es_nocturno = False
        if emp_turno.get('hora_salida') and emp_turno.get('hora_ingreso'):
            es_nocturno = emp_turno['hora_salida'] < emp_turno['hora_ingreso']

        if es_nocturno:
            # Ventana Nocturna: Desde hoy al mediodía hasta mañana al mediodía
            inicio_ventana = f"{fecha} 12:00:00"
            fin_ventana = f"{fecha + timedelta(days=1)} 11:59:59"
        else:
            # Ventana Diurna: El día calendario normal
            inicio_ventana = f"{fecha} 00:00:00"
            fin_ventana = f"{fecha} 23:59:59"

        # 📡 Extraer los Marcajes Brutos usando la ventana de tiempo
        cur.execute(f"""
            SELECT fecha_hora FROM {schema}.eventos_brutos 
            WHERE item = %s AND fecha_hora >= %s AND fecha_hora <= %s 
            ORDER BY fecha_hora ASC
        """, (bio_id_str, inicio_ventana, fin_ventana))
        marcajes = [m['fecha_hora'] for m in cur.fetchall()]

        # 3. Obtener Permisos/Ausencias para ese día
        cur.execute(f"""
            SELECT tipo, hora_inicio, hora_fin 
            FROM {schema}.ausencias 
            WHERE empleado_id = %s AND estado = 'aprobado' AND eliminado = FALSE
            AND %s BETWEEN fecha_inicio AND fecha_fin
        """, (empleado_id, fecha))
        ausencias = cur.fetchall()

        # 🚀 4. EJECUTAR MOTOR MATEMÁTICO
        resumen = calcular_dia_asistencia(marcajes, emp_turno, ausencias, emp_turno['salario_base'], fecha)

        # 5. UPSERT (Insertar o Actualizar) en la tabla Caché
        cur.execute(f"""
            INSERT INTO {schema}.asistencia_diaria 
            (empleado_id, fecha, turno_id, hora_entrada, hora_inicio_almuerzo, hora_fin_almuerzo, hora_salida, 
             minutos_retraso_entrada, minutos_exceso_almuerzo, horas_trabajadas, horas_extras, estado, deuda_generada_bs, actualizado_en)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (empleado_id, fecha) 
            DO UPDATE SET 
                hora_entrada = EXCLUDED.hora_entrada,
                hora_inicio_almuerzo = EXCLUDED.hora_inicio_almuerzo,
                hora_fin_almuerzo = EXCLUDED.hora_fin_almuerzo,
                hora_salida = EXCLUDED.hora_salida,
                minutos_retraso_entrada = EXCLUDED.minutos_retraso_entrada,
                minutos_exceso_almuerzo = EXCLUDED.minutos_exceso_almuerzo,
                horas_trabajadas = EXCLUDED.horas_trabajadas,
                horas_extras = EXCLUDED.horas_extras,
                estado = EXCLUDED.estado,
                deuda_generada_bs = EXCLUDED.deuda_generada_bs,
                actualizado_en = CURRENT_TIMESTAMP
            WHERE {schema}.asistencia_diaria.modificado_manualmente = FALSE;
        """, (
            empleado_id, fecha, emp_turno['turno_id'], resumen['hora_entrada'], 
            resumen['hora_inicio_almuerzo'], resumen['hora_fin_almuerzo'], resumen['hora_salida'], 
            resumen['minutos_retraso_entrada'], resumen['minutos_exceso_almuerzo'], 
            resumen['horas_trabajadas'], resumen.get('horas_extras', 0.0), resumen['estado'], resumen['deuda_generada_bs']
        ))
        
        conn.commit()
        return True
    except Exception as e:
        print(f"❌ Error procesando día: {e}")
        conn.rollback()
        return False
    finally:
        cur.close()
        conn.close()


# ==============================================================================
# 5. RUTAS PARA COMUNICACIÓN CON HARDWARE (ZKTeco ADMS)
# ==============================================================================

@app.get("/iclock/cdata")
async def iclock_init(request: Request):
    sn = request.query_params.get("SN", "")
    print(f"✅ Lector intentando conectar: SN={sn}")
    return PlainTextResponse(f"GET OPTION FROM: {sn}\nATTLOGStamp=None\nOPERLOGStamp=9999\nRealtime=1\nEncrypt=None\n")

# ── EL RELOJ PREGUNTA SI HAY ÓRDENES PENDIENTES ──
@app.get("/iclock/getrequest")
async def iclock_getrequest(request: Request):
    sn = request.query_params.get("SN", "")
    # Le respondemos "OK" con salto de línea para indicarle que NO hay comandos pendientes.
    # Así el reloj se queda tranquilo y deja de hacer spam.
    return PlainTextResponse("OK\n")

# ── RECEPCIÓN DE MARCAJES DEL LECTOR (VERSIÓN ASÍNCRONA ULTRA-RÁPIDA) ──
@app.post("/iclock/cdata")
async def iclock_data(request: Request, background_tasks: BackgroundTasks):
    table = request.query_params.get("table", "")
    sn = request.query_params.get("SN", "")
    body = await request.body()
    texto = body.decode("utf-8", errors="ignore")
    
    if table == "ATTLOG":
        # 1. ENRUTADOR GLOBAL
        conn_maestra = conectar_bd("public")
        cur_m = conn_maestra.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        try:
            cur_m.execute("SELECT schema_name FROM public.dispositivos WHERE numero_serie = %s AND activo = TRUE", (sn,))
            disp = cur_m.fetchone()
            
            if not disp:
                print(f"⚠️ Reloj desconocido (SN: {sn}).")
                return PlainTextResponse("OK\n")
                
            schema_destino = disp["schema_name"]
            
            # 2. Procesamos las huellas
            conn_e = conectar_bd(schema_destino)
            cur_e = conn_e.cursor()

            for linea in texto.strip().splitlines():
                partes = linea.strip().split("\t")
                if len(partes) >= 2:
                    bio_id = partes[0]
                    fecha_hora_str = partes[1] 
                    fecha_dt = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M:%S").date()
                    
                    # ⚡ ESCUDO ANTI-DUPLICADOS: Si la huella exacta ya está, la ignoramos
                    cur_e.execute(f"SELECT id FROM {schema_destino}.eventos_brutos WHERE device_no = %s AND item = %s AND fecha_hora = %s", (sn, bio_id, fecha_hora_str))
                    if cur_e.fetchone():
                        continue 
                        
                    # A) Guardamos en la Caja Negra
                    cur_e.execute(f"""
                        INSERT INTO {schema_destino}.eventos_brutos (device_no, item, verify_mode, action, fecha_hora, raw_data)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (sn, bio_id, partes[3] if len(partes)>3 else "1", partes[2] if len(partes)>2 else "0", fecha_hora_str, psycopg2.extras.Json({"raw": linea})))

                    # B) ⚡ ENCOLAMOS EL CÁLCULO EN SEGUNDO PLANO (El servidor ya no se congela)
                    cur_e.execute(f"SELECT id FROM {schema_destino}.empleados WHERE bio_id = %s", (bio_id,))
                    res_emp = cur_e.fetchone()
                    if res_emp:
                        background_tasks.add_task(procesar_asistencia_dia, schema_destino, res_emp[0], fecha_dt)
                        print(f"🚀 Marcaje encolado en [{schema_destino}] para BioID: {bio_id}")

            conn_e.commit()
            cur_e.close()
            conn_e.close()

        except Exception as e:
            print(f"❌ Error interno ADMS: {e}")
        finally:
            cur_m.close()
            conn_maestra.close()
            
    # ⚡ FIX FINAL: El salto de línea \n es OBLIGATORIO para que el ZKTeco borre su memoria y deje de enviar
    return PlainTextResponse("OK\n")


# ==============================================================================
# 6. RUTA PARA ELIMINAR UNA EMPRESA (SOLO SUPERADMIN) ──
# ==============================================================================

@app.delete("/empresas/{empresa_id}")
def eliminar_empresa(empresa_id: int, usuario = Depends(verificar_token)):
    # 1. SEGURIDAD: Verificamos usando tu propia función de seguridad
    if usuario.get("rol") != "superadmin":
        raise HTTPException(status_code=403, detail="Acceso denegado. Solo SuperAdmin puede eliminar empresas.")

    # 2. Conectamos a la base de datos maestra (public)
    conn = conectar_bd("public")
    cur = conn.cursor()
    
    try:
        # 3. BUSCAR EL OBJETIVO: Obtenemos el nombre del esquema (carpeta) de esa empresa
        cur.execute("SELECT schema_name FROM empresas WHERE id = %s", (empresa_id,))
        empresa = cur.fetchone()
        
        if not empresa:
            raise HTTPException(status_code=404, detail="La empresa no existe.")
            
        schema_name = empresa[0] # Extraemos el texto, ej: "empresa_industrias_prueba"

        # 4. LA EXPLOSIÓN CONTROLADA (DROP SCHEMA CASCADE)
        from psycopg2 import sql
        cur.execute(sql.SQL("DROP SCHEMA IF EXISTS {} CASCADE").format(sql.Identifier(schema_name)))
        
        # 5. LIMPIEZA DE RASTROS
        cur.execute("DELETE FROM usuarios WHERE empresa_id = %s", (empresa_id,))
        cur.execute("DELETE FROM empresas WHERE id = %s", (empresa_id,))
        
        # 6. CONFIRMAR CAMBIOS
        conn.commit()
        
        return {"mensaje": f"La empresa y toda su información fueron eliminadas de raíz."}
        
    except Exception as e:
        conn.rollback() # Si algo falla, cancelamos la explosión
        raise HTTPException(status_code=500, detail=f"Error al eliminar: {str(e)}")
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 7. RUTAS DE CATÁLOGOS (ESTRUCTURA ORGANIZACIONAL Y TIEMPO)
# ==============================================================================

# ------------------------------------------------------------------------------
# A. SUCURSALES (Físico)
# ------------------------------------------------------------------------------
@app.get("/sucursales")
def obtener_sucursales(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {schema}.sucursales WHERE eliminado = FALSE ORDER BY nombre")
    sucursales = cur.fetchall()
    cur.close()
    conn.close()
    return list(sucursales)

@app.post("/sucursales")
async def crear_sucursal(data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # ⚡ CORRECCIÓN: 4 campos (%s) = 4 variables en la tupla inferior
        cur.execute(f"""
            INSERT INTO {schema}.sucursales (nombre, ciudad, direccion, telefono) 
            VALUES (%s, %s, %s, %s)
        """, (
            data.get("nombre"), 
            data.get("ciudad"),       # ⚡ ¡ESTA ES LA LÍNEA QUE FALTABA!
            data.get("direccion"), 
            data.get("telefono", "") 
        ))
        conn.commit()
        return {"mensaje": "Sucursal creada exitosamente"}
    except Exception as e:
        conn.rollback()
        print(f"ERROR CRÍTICO EN POST SUCURSALES: {e}") 
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

# --- RUTA PARA ACTUALIZAR SUCURSAL (EDICIÓN) ---
@app.put("/sucursales/{sucursal_id}")
async def actualizar_sucursal(sucursal_id: int, data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # Ejecutamos la actualización de los 3 campos principales
        cur.execute(f"""
            UPDATE {schema}.sucursales 
            SET nombre = %s, ciudad = %s, direccion = %s, telefono = %s 
            WHERE id = %s
        """, (
            data.get("nombre"),
            data.get("ciudad"),  
            data.get("direccion"), 
            data.get("telefono", ""), 
            sucursal_id
        ))
        
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Sucursal no encontrada")
            
        conn.commit()
        return {"mensaje": "Sucursal actualizada correctamente"}
    except Exception as e:
        conn.rollback()
        print(f"Error al actualizar sucursal: {e}")
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/sucursales/{sucursal_id}")
def eliminar_sucursal(sucursal_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # Soft Delete: En lugar de borrar la fila, la ocultamos de la interfaz
        cur.execute(f"UPDATE {schema}.sucursales SET eliminado = TRUE WHERE id = %s", (sucursal_id,))
        conn.commit()
        return {"mensaje": "Sucursal eliminada correctamente"}
    except psycopg2.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Operación denegada: No puedes eliminar esta sucursal porque existen empleados registrados (activos o inactivos) asignados a ella. Transfiérelos a otra sucursal primero.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# B. SECCIONES (Lógico)
# ------------------------------------------------------------------------------
@app.get("/secciones")
def obtener_secciones(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"SELECT * FROM {schema}.secciones WHERE eliminado = FALSE ORDER BY nombre")
    secciones = cur.fetchall()
    cur.close()
    conn.close()
    return list(secciones)

@app.post("/secciones")
async def crear_seccion(request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    data = await request.json()
    nombre = data.get("nombre")
    descripcion = data.get("descripcion", "")
    
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        cur.execute(
            f"INSERT INTO {schema}.secciones (nombre, descripcion) VALUES (%s, %s) RETURNING id",
            (nombre, descripcion)
        )
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return {"mensaje": "Sección registrada con éxito", "id": nuevo_id}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# --- RUTA PARA ACTUALIZAR SECCIÓN (EDICIÓN) ---
@app.put("/secciones/{seccion_id}")
async def actualizar_seccion(seccion_id: int, data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        cur.execute(f"""
            UPDATE {schema}.secciones 
            SET nombre = %s, descripcion = %s 
            WHERE id = %s
        """, (
            data.get("nombre"), 
            data.get("descripcion"), 
            seccion_id
        ))
        
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Sección no encontrada")
            
        conn.commit()
        return {"mensaje": "Sección actualizada correctamente"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/secciones/{seccion_id}")
def eliminar_seccion(seccion_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE {schema}.secciones SET eliminado = TRUE WHERE id = %s", (seccion_id,))
        conn.commit()
        return {"mensaje": "Sección eliminada correctamente"}
    except psycopg2.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Operación denegada: No puedes eliminar esta sección (departamento) porque tienes personal asignado a la misma. Reasigna al personal primero.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 8. GESTIÓN DE PLANILLA (EMPLEADOS)
# ==============================================================================

# ⚡ NUEVO ENDPOINT: Extrae estadísticas y cargos únicos en milisegundos
@app.get("/empleados/stats")
async def obtener_empleados_stats(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # 1. Contamos usando la velocidad nativa de PostgreSQL
        cur.execute(f"""
            SELECT 
                COUNT(*) filter (where activo = true) as activos, 
                COUNT(*) filter (where activo = false) as inactivos, 
                COUNT(*) as todos 
            FROM {schema}.empleados WHERE eliminado = FALSE
        """)
        counts = cur.fetchone()

        # 2. Extraemos los cargos únicos sin duplicados
        cur.execute(f"""
            SELECT DISTINCT cargo 
            FROM {schema}.empleados 
            WHERE eliminado = FALSE AND cargo IS NOT NULL AND cargo != '' 
            ORDER BY cargo
        """)
        cargos = [row['cargo'] for row in cur.fetchall()]

        return {
            "activos": counts['activos'] or 0,
            "inactivos": counts['inactivos'] or 0,
            "todos": counts['todos'] or 0,
            "cargos": cargos
        }
    finally:
        cur.close()
        conn.close()


# ⚡ ENDPOINT ACTUALIZADO: Buscador avanzado (Dumb Frontend)
@app.get("/empleados")
async def obtener_empleados(
    estado: str = 'activos', 
    q: str = '', 
    sucursal_id: str = '', 
    seccion_id: str = '', 
    cargo: str = '',
    limite: int = 500,    # Límite alto por defecto para permitir exportar a Excel
    offset: int = 0,
    usuario = Depends(verificar_token)
):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # 1. Armamos las condiciones dinámicamente según lo que pida el Frontend Tonto
        condiciones = ["e.eliminado = FALSE"]
        parametros = []
        
        if estado == 'activos': condiciones.append("e.activo = TRUE")
        elif estado == 'inactivos': condiciones.append("e.activo = FALSE")
            
        if q:
            condiciones.append("(e.nombres ILIKE %s OR e.apellidos ILIKE %s OR e.ci ILIKE %s OR e.bio_id::text = %s)")
            param_q = f"%{q}%"
            parametros.extend([param_q, param_q, param_q, q])
            
        if sucursal_id:
            condiciones.append("e.sucursal_id = %s")
            parametros.append(int(sucursal_id))
            
        if seccion_id:
            condiciones.append("e.seccion_id = %s")
            parametros.append(int(seccion_id))
            
        if cargo:
            condiciones.append("e.cargo = %s")
            parametros.append(cargo)
            
        where_clause = " AND ".join(condiciones)
        
        query = f"""
            SELECT e.*, 
                   s.nombre as sucursal_nombre,
                   s.ciudad as sucursal_ciudad,
                   sec.nombre as seccion_nombre,
                   t.nombre as turno_nombre,
                   t.hora_ingreso as turno_ingreso,
                   t.hora_salida as turno_salida,
                   t.almuerzo as turno_almuerzo,
                   t.almuerzo_min as turno_almuerzo_min,
                   (SELECT tipo 
                    FROM {schema}.ausencias a 
                    WHERE a.empleado_id = e.id 
                      AND a.estado = 'aprobado' 
                      AND a.eliminado = FALSE 
                      AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin 
                    LIMIT 1) as estado_ausencia
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE {where_clause}
            ORDER BY e.activo DESC, e.nombres ASC, e.apellidos ASC
            LIMIT %s OFFSET %s
        """
        parametros.extend([limite, offset])
        
        cur.execute(query, parametros)
        empleados = cur.fetchall()
        
        # 2. Formateo seguro para evitar que el JSON explote
        for emp in empleados:
            for key in ["fecha_ingreso", "fecha_antiguedad", "fecha_retiro", "turno_ingreso", "turno_salida", "creado_en", "actualizado_en"]:
                if emp.get(key): emp[key] = str(emp[key])
                
        return empleados
    finally:
        cur.close()
        conn.close()

@app.post("/empleados")
async def crear_empleado(request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    data = await request.json()
    
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        ci = data.get("ci")

        # ⚡ Lógica de ID Inteligente (Híbrida)
        bio_id_recibido = data.get('bio_id')
        
        # Si el campo está vacío o es None, autogeneramos
        if bio_id_recibido is None or str(bio_id_recibido).strip() == "":
            cur.execute(f"SELECT MAX(bio_id) FROM {schema}.empleados")
            res_max = cur.fetchone()
            
            # ⚡ CORRECCIÓN CLAVE: Leemos la posición [0] de la tupla
            max_actual = res_max[0] if res_max and res_max[0] is not None else 0
            bio_id_final = max_actual + 1
        else:
            bio_id_final = int(bio_id_recibido)

        # Verificamos que no exista un duplicado
        cur.execute(f"SELECT id FROM {schema}.empleados WHERE bio_id = %s", (bio_id_final,))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail=f"El ID de lector {bio_id_final} ya está asignado a otro empleado.")
        turno_id_raw = data.get("turno_id")
        turno_id_final = int(turno_id_raw) if turno_id_raw else None
        
        # ⚡ Capturamos la foto Base64 enviada por el frontend
        foto_perfil = data.get("foto_perfil") 

        cur.execute(f"SELECT id, eliminado FROM {schema}.empleados WHERE ci = %s", (ci,))
        existe = cur.fetchone()

        if existe:
            id_db, esta_eliminado = existe
            if not esta_eliminado:
                # ⚡ ALERTA MEJORADA (409)
                raise HTTPException(status_code=409, detail="ERROR: Ese C.I. ya existe. Si fue retirado, búscalo en 'Inactivos' y cámbialo a 'Activo'.")
            else:
                cur.execute(f"""
                    UPDATE {schema}.empleados 
                    SET bio_id = %s, foto_perfil = %s, nombres = %s, apellidos = %s, 
                        sucursal_id = %s, seccion_id = %s, cargo = %s, turno_id = %s,
                        eliminado = FALSE, activo = TRUE
                    WHERE id = %s
                """, (bio_id_final, foto_perfil, data.get("nombres"), data.get("apellidos"), 
                      data.get("sucursal_id"), data.get("seccion_id"), data.get("cargo"), turno_id_final, id_db))
                msg = "Empleado reactivado correctamente."
        else:
            cur.execute(f"""
            INSERT INTO {schema}.empleados 
            (bio_id, foto_perfil, nombres, apellidos, ci, sucursal_id, seccion_id, cargo, turno_id, activo,
             sexo, celular, correo, direccion, fecha_ingreso, fecha_antiguedad, tipo_contrato, salario_base, bono) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            bio_id_final, 
            foto_perfil,          # ⚡ Inyectamos la foto aquí
            data.get("nombres"), 
            data.get("apellidos"), 
            ci, 
            data.get("sucursal_id"), 
            data.get("seccion_id"), 
            data.get("cargo"), 
            turno_id_final,       
            True,                 
            data.get("sexo"), 
            data.get("celular"), 
            data.get("correo"), 
            data.get("direccion"),
            data.get("fecha_ingreso"), 
            data.get("fecha_antiguedad"), 
            data.get("tipo_contrato"),
            data.get("salario_base", 0), 
            data.get("bono", 0)   
        ))
            msg = "Personal registrado con éxito."

        conn.commit()
        return {"mensaje": msg}
        
    except psycopg2.IntegrityError:
        conn.rollback()
        # ⚡ ALERTA MEJORADA (409)
        raise HTTPException(status_code=409, detail="El ID Biométrico ya está ocupado por otro empleado activo.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.put("/empleados/{empleado_id}")
async def actualizar_empleado(empleado_id: int, request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    data = await request.json()
    
    # --- 1. SEGURIDAD ESTRICTA: VERIFICAR CONTRASEÑA DE BAJA ---
    activo_nuevo = data.get("activo", True)
    
    if not activo_nuevo: # Si la orden es "Apagar" al empleado
        admin_password = data.get("admin_password")
        if not admin_password:
            raise HTTPException(status_code=409, detail="Contraseña de administrador requerida para ejecutar la baja.")
        
        # Conectamos a la base maestra para ver la contraseña del usuario que está intentando hacer esto
        conn_pub = conectar_bd("public")
        cur_pub = conn_pub.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur_pub.execute("SELECT password_hash FROM usuarios WHERE id = %s", (usuario["id"],))
        user_db = cur_pub.fetchone()
        cur_pub.close()
        conn_pub.close()

        # Comparamos la contraseña digitada contra el Hash de la Base de Datos
        if not user_db or not bcrypt.checkpw(admin_password.encode(), user_db["password_hash"].encode()):
            raise HTTPException(status_code=409, detail="Contraseña incorrecta. Operación de baja denegada.")

    # --- 2. ACTUALIZACIÓN Y CAJA NEGRA ---
    conn = conectar_bd(schema)
    cur = conn.cursor()
    # --- 2. ACTUALIZACIÓN Y CAJA NEGRA ---
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # A. Consultamos el estado ACTUAL antes de guardar los cambios
        cur.execute(f"SELECT activo, fecha_retiro, motivo_retiro, historial_movimientos, foto_perfil FROM {schema}.empleados WHERE id = %s", (empleado_id,))
        estado_db = cur.fetchone()
        
        activo_actual = estado_db[0]
        historial_existente = estado_db[3] or ""
        foto_actual_bd = estado_db[4] # Guardamos la foto que ya existe
        nuevo_historial = historial_existente
        
        fecha_retiro_final = data.get("fecha_retiro")
        motivo_retiro_final = data.get("motivo_retiro")
        
        fecha_auditoria = datetime.now().strftime("%Y-%m-%d")

        # B. LÓGICA DE AUDITORÍA (CAJA NEGRA)
        if not activo_actual and activo_nuevo:
            # CASO: RECONTRATACIÓN
            old_fecha = estado_db[1] or "Desconocida"
            old_motivo = estado_db[2] or "Sin motivo especificado"
            if old_fecha != "Desconocida":
                anotacion = f"RECONTRATADO EL {fecha_auditoria}: Su baja anterior fue el {old_fecha} por '{old_motivo}'.\n"
                nuevo_historial = anotacion + historial_existente
            
            fecha_retiro_final = None
            motivo_retiro_final = None

        elif activo_actual and not activo_nuevo:
            # CASO: DESPIDO / RETIRO
            anotacion = f"➤ [DADO DE BAJA EL {fecha_auditoria}]: Motivo registrado: {motivo_retiro_final}.\n"
            nuevo_historial = anotacion + historial_existente

        # ⚡ BLINDAJE DE FOTO UNIFICADO (El embudo perfecto)
        foto_perfil = data.get("foto_perfil")
        
        if foto_perfil == "ELIMINAR":
            foto_a_guardar = None  # 1. El usuario hizo clic en la 'X' roja, limpiamos la foto
        elif foto_perfil:
            foto_a_guardar = foto_perfil  # 2. El usuario subió una cámara/archivo nuevo
        else:
            foto_a_guardar = foto_actual_bd  # 3. El usuario no tocó nada, conservamos la que ya estaba en BD

        # C. GUARDAMOS TODO EN POSTGRESQL EN UNA SOLA PETICIÓN
        cur.execute(f"""
            UPDATE {schema}.empleados 
            SET bio_id = %s, nombres = %s, apellidos = %s, ci = %s, 
                sucursal_id = %s, seccion_id = %s, cargo = %s, activo = %s,
                sexo = %s, celular = %s, correo = %s, direccion = %s,
                fecha_ingreso = %s, fecha_antiguedad = %s, tipo_contrato = %s, 
                salario_base = %s, bono = %s,
                fecha_retiro = %s, motivo_retiro = %s,
                turno_id = %s,
                historial_movimientos = %s,
                foto_perfil = %s
            WHERE id = %s
        """, (
            data.get("bio_id"), data.get("nombres"), data.get("apellidos"), 
            data.get("ci"), data.get("sucursal_id"), data.get("seccion_id"), 
            data.get("cargo"), activo_nuevo,
            data.get("sexo"), data.get("celular"), data.get("correo"), data.get("direccion"),
            data.get("fecha_ingreso"), data.get("fecha_antiguedad"), data.get("tipo_contrato"),
            data.get("salario_base", 0), data.get("bono", 0),
            fecha_retiro_final, motivo_retiro_final,
            data.get("turno_id"),
            nuevo_historial,
            foto_a_guardar, # ⚡ Aquí entra nuestra variable ya definida y procesada
            empleado_id
        ))
        
        conn.commit()
        return {"mensaje": "Perfil actualizado y bitácora registrada con éxito."}
        
    except psycopg2.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="El C.I. o ID Biométrico ya está siendo usado por otro empleado en el sistema.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/empleados/{empleado_id}")
async def eliminar_empleado(empleado_id: int, request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # TRUCO SENIOR: Renombramos el CI para liberarlo y ponemos bio_id en NULL
        cur.execute(f"""
            UPDATE {schema}.empleados 
            SET eliminado = TRUE, 
                activo = FALSE, 
                bio_id = NULL,
                ci = ci || '-DEL-' || id  -- Agrega "-DEL-ID" al carnet para liberar el número original
            WHERE id = %s
        """, (empleado_id,))
        
        conn.commit()
        return {"mensaje": "Registro de prueba eliminado. C.I. y ID Biométrico liberados."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ⚡ NUEVO: FABRICANTE DE EXCEL SERVER-SIDE
@app.get("/empleados/exportar/excel")
async def exportar_empleados_excel(estado: str="activos", q: str="", sucursal_id: str="", seccion_id: str="", cargo: str="", usuario=Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # ⚡ CONSULTA SQL COMPLETA (Sin atajos)
        query = f"""
            SELECT e.*, s.nombre as sucursal_nombre, sec.nombre as seccion_nombre, t.nombre as turno_nombre
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.eliminado = FALSE
        """
        parametros = []
        if estado == "activos": query += " AND e.activo = TRUE"
        elif estado == "inactivos": query += " AND e.activo = FALSE"
        
        if sucursal_id and sucursal_id.isdigit(): 
            query += " AND e.sucursal_id = %s"
            parametros.append(int(sucursal_id))
            
        if seccion_id and seccion_id.isdigit(): 
            query += " AND e.seccion_id = %s"
            parametros.append(int(seccion_id))
            
        if cargo: 
            query += " AND e.cargo = %s"
            parametros.append(cargo)
            
        if q:
            query += " AND (e.nombres ILIKE %s OR e.apellidos ILIKE %s OR e.ci ILIKE %s OR CAST(e.bio_id AS TEXT) ILIKE %s)"
            termino = f"%{q}%"
            parametros.extend([termino, termino, termino, termino])
        
        cur.execute(query + " ORDER BY e.nombres ASC", tuple(parametros))
        empleados = cur.fetchall()

        # ⚡ CREACIÓN DEL EXCEL EN MEMORIA
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilla_Personal"

        # 1. Cabecera Exacta de 20 columnas
        headers = [
            "ID Lector", "Nombres", "Apellidos", "C.I.", "Sexo", "Celular", "Correo", 
            "Cargo", "Sucursal", "Sección", "Turno", "Fecha Ingreso", "Inicio Antigüedad", "Contrato", 
            "Salario Base", "Bono", "Estado", "Fecha Baja", "Motivo Baja", "Historial de Movimientos"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 2. Inyección de Datos Segura
        for emp in empleados:
            ws.append([
                emp.get('bio_id') or '', 
                emp.get('nombres') or '', 
                emp.get('apellidos') or '', 
                emp.get('ci') or '',
                emp.get('sexo') or '', 
                emp.get('celular') or '', 
                emp.get('correo') or '',
                emp.get('cargo') or '', 
                emp.get('sucursal_nombre') or '', 
                emp.get('seccion_nombre') or '',
                emp.get('turno_nombre') or 'Sin Turno',
                str(emp['fecha_ingreso']).split(' ')[0] if emp.get('fecha_ingreso') else '',
                str(emp['fecha_antiguedad']).split(' ')[0] if emp.get('fecha_antiguedad') else '',
                emp.get('tipo_contrato') or '',
                float(emp.get('salario_base') or 0), 
                float(emp.get('bono') or 0),
                'ACTIVO' if emp.get('activo') else 'INACTIVO',
                str(emp['fecha_retiro']).split(' ')[0] if emp.get('fecha_retiro') else '',
                emp.get('motivo_retiro') or '', 
                emp.get('historial_movimientos') or ''
            ])

        # 3. Ajuste de anchos respetando tu diseño original
        anchos = [10, 20, 20, 15, 12, 15, 25, 20, 15, 15, 15, 18, 15, 12, 10, 10, 15, 18, 30, 50]
        for i, ancho in enumerate(anchos, 1):
            # ⚡ FIX: Llamamos a la ruta completa para evitar errores de importación
            col_letter = openpyxl.utils.get_column_letter(i) 
            ws.column_dimensions[col_letter].width = ancho

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename=Planilla_{estado}.xlsx"}
        )
    finally:
        cur.close()
        conn.close()

# ⚡ NUEVO: FABRICANTE DE PDF SERVER-SIDE
@app.get("/empleados/exportar/pdf")
async def exportar_empleados_pdf(estado: str="activos", q: str="", sucursal_id: str="", seccion_id: str="", cargo: str="", usuario=Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # 1. CONSTRUCCIÓN DE LA ETIQUETA DE FILTROS (Inteligencia de Negocio)
        filtros_texto = f"Estado: {estado.upper()}"
        
        if sucursal_id:
            cur.execute(f"SELECT nombre FROM {schema}.sucursales WHERE id = %s", (sucursal_id,))
            res = cur.fetchone()
            if res: filtros_texto += f" | Sucursal: {res['nombre']}"
            
        if seccion_id:
            cur.execute(f"SELECT nombre FROM {schema}.secciones WHERE id = %s", (seccion_id,))
            res = cur.fetchone()
            if res: filtros_texto += f" | Sección: {res['nombre']}"
            
        if cargo:
            filtros_texto += f" | Cargo: {cargo}"
            
        if q:
            filtros_texto += f" | Búsqueda: '{q}'"
        
        # ⚡ RECONSTRUIDO: La consulta SQL completa y real (nada de tres puntos)
        query = f"""
            SELECT e.*, s.nombre as sucursal_nombre, sec.nombre as seccion_nombre, t.nombre as turno_nombre
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.eliminado = FALSE
        """
        parametros = []
        if estado == "activos": query += " AND e.activo = TRUE"
        elif estado == "inactivos": query += " AND e.activo = FALSE"
        if sucursal_id and sucursal_id.isdigit(): query += " AND e.sucursal_id = %s"; parametros.append(int(sucursal_id))
        if seccion_id and seccion_id.isdigit(): query += " AND e.seccion_id = %s"; parametros.append(int(seccion_id))
        if cargo: query += " AND e.cargo = %s"; parametros.append(cargo)
        if q:
            query += " AND (e.nombres ILIKE %s OR e.apellidos ILIKE %s OR e.ci ILIKE %s OR CAST(e.bio_id AS TEXT) ILIKE %s)"
            termino = f"%{q}%"
            parametros.extend([termino, termino, termino, termino])
        
        cur.execute(query + " ORDER BY e.nombres ASC", tuple(parametros))
        empleados = cur.fetchall()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        style_bloque = styles['Normal']
        style_bloque.fontSize = 7
        style_bloque.leading = 9 

        elements.append(Paragraph(f"<b>Reporte Maestro de Planilla - Empleados {estado.upper()}</b>", styles['Title']))
        from datetime import datetime # Nos aseguramos de tener la hora
        elements.append(Paragraph(f"<font size=10 color=gray>Filtro: {estado.upper()} | Emisión: {datetime.now().strftime('%d/%m/%Y')}</font>", styles['Normal']))
        elements.append(Paragraph("<br/><br/>", styles['Normal']))

        data = [["ID", "Datos Personales", "Contacto", "Ubicación Laboral", "Contrato e Ingresos", "Estado / Retiro"]]
        
        for emp in empleados:
            personal = f"<b>{emp['nombres']} {emp['apellidos']}</b><br/>C.I.: {emp['ci']}<br/>Sexo: {emp.get('sexo') or 'N/A'}"
            contacto = f"Cel: {emp.get('celular') or '-'}<br/>Correo: {emp.get('correo') or '-'}<br/>Dir: {emp.get('direccion') or '-'}"
            laboral = f"Cargo: {emp.get('cargo') or '-'}<br/>Sucursal: {emp.get('sucursal_nombre') or '-'}<br/>Sección: {emp.get('seccion_nombre') or '-'}<br/>Turno: {emp.get('turno_nombre') or 'No asignado'}<br/>Ingreso: {str(emp['fecha_ingreso']).split(' ')[0] if emp.get('fecha_ingreso') else '-'}"
            ingresos = f"Contrato: {emp.get('tipo_contrato') or '-'}<br/>Salario: Bs. {emp.get('salario_base') or '0.00'}<br/>Bono: Bs. {emp.get('bono') or '0.00'}<br/>Inicio Ant.: {str(emp['fecha_antiguedad']).split(' ')[0] if emp.get('fecha_antiguedad') else '-'}"
            
            retiro_txt = "<b>ACTIVO</b>" if emp['activo'] else "<b>INACTIVO</b>"
            if not emp['activo']:
                retiro_txt += f"<br/>Fecha Baja: {str(emp['fecha_retiro']).split(' ')[0] if emp.get('fecha_retiro') else '-'}<br/>Motivo: {emp.get('motivo_retiro') or 'No especificado'}"
            if emp.get('historial_movimientos'):
                retiro_txt += f"<br/><br/><i>-- HISTORIAL --</i><br/>{emp['historial_movimientos']}"

            data.append([
                str(emp.get('bio_id') or '-'),
                Paragraph(personal, style_bloque),
                Paragraph(contacto, style_bloque),
                Paragraph(laboral, style_bloque),
                Paragraph(ingresos, style_bloque),
                Paragraph(retiro_txt, style_bloque)
            ])

        t = Table(data, colWidths=[40, 130, 130, 140, 130, 140])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
        ]))
        elements.append(t)
        doc.build(elements)
        output.seek(0)
        return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=Planilla_{estado}.pdf"})
    finally:
        cur.close()
        conn.close()

# ==========================================
# 9. MÓDULO: TURNOS Y HORARIOS
# ==========================================

@app.get("/turnos")
async def obtener_turnos(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(f"SELECT * FROM {schema}.turnos WHERE eliminado = FALSE ORDER BY id ASC")
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()

def calcular_minutos_almuerzo(inicio: str, fin: str) -> int:
    """Función auxiliar para calcular la duración del almuerzo en minutos"""
    if not inicio or not fin:
        return 0
    try:
        t_ini = datetime.strptime(inicio, "%H:%M")
        t_fin = datetime.strptime(fin, "%H:%M")
        # Si el fin es menor al inicio, asumimos cruce de medianoche
        if t_fin < t_ini:
            t_fin += timedelta(days=1)
        return int((t_fin - t_ini).total_seconds() / 60)
    except:
        return 0

# ==========================================
# CALCULADORA INTELIGENTE DE TURNOS
# ==========================================
@app.get("/turnos/calculadora")
async def calculadora_turnos(ingreso: str = "", salida: str = "", alm_in: str = "", alm_out: str = "", usuario = Depends(verificar_token)):
    """Calculadora centralizada para mantener el Frontend Tonto"""
    min_almuerzo = 0
    total_str = "0.00 hrs/día"
    try:
        if alm_in and alm_out:
            h1, m1 = map(int, alm_in.split(':'))
            h2, m2 = map(int, alm_out.split(':'))
            m_total = ((h2 * 60) + m2) - ((h1 * 60) + m1)
            if m_total < 0: m_total += 1440
            min_almuerzo = m_total

        if ingreso and salida:
            h1, m1 = map(int, ingreso.split(':'))
            h2, m2 = map(int, salida.split(':'))
            m_total = ((h2 * 60) + m2) - ((h1 * 60) + m1)
            if m_total < 0: m_total += 1440
            
            min_netos = m_total - min_almuerzo
            total_str = f"{(max(0, min_netos) / 60):.2f} hrs/día"
    except: pass
    return {"almuerzo_min": min_almuerzo, "total_str": total_str}

@app.post("/turnos")
async def crear_turno(data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # ⚡ Mapeo exacto con la DB y el JSON del frontend
        cur.execute(f"""
            INSERT INTO {schema}.turnos 
            (nombre, hora_ingreso, hora_salida, dias, almuerzo, hora_inicio_almuerzo, hora_fin_almuerzo, almuerzo_min, tolerancia_min)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['nombre'], 
            data['hora_ingreso'], 
            data['hora_salida'], 
            psycopg2.extras.Json(data.get('dias', [])), # ⚡ Corrección del error 'dias' NOT NULL
            data.get('almuerzo', True),
            data.get('inicio_almuerzo'), 
            data.get('fin_almuerzo'), 
            data.get('almuerzo_min', 0),
            data.get('tolerancia_min', 0)
        ))
        conn.commit()
        return {"mensaje": "Turno creado exitosamente"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en DB: {str(e)}")
    finally:
        cur.close(); conn.close()

@app.put("/turnos/{turno_id}")
async def actualizar_turno(turno_id: int, data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    
    # ⚡ El Cerebro calcula antes de guardar
    almuerzo_min = 0
    if data.get('inicio_almuerzo') and data.get('fin_almuerzo'):
        h1, m1 = map(int, data['inicio_almuerzo'].split(':'))
        h2, m2 = map(int, data['fin_almuerzo'].split(':'))
        almuerzo_min = ((h2 * 60) + m2) - ((h1 * 60) + m1)
        if almuerzo_min < 0: almuerzo_min += 1440

    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # ⚡ UPDATE con todas las columnas sincronizadas
        cur.execute(f"""
            UPDATE {schema}.turnos 
            SET nombre = %s, 
                hora_ingreso = %s, 
                hora_salida = %s, 
                dias = %s, 
                almuerzo = %s, 
                hora_inicio_almuerzo = %s, 
                hora_fin_almuerzo = %s, 
                almuerzo_min = %s, 
                tolerancia_min = %s
            WHERE id = %s
        """, (
            data['nombre'], 
            data['hora_ingreso'], 
            data['hora_salida'], 
            psycopg2.extras.Json(data.get('dias', [])), # ⚡ Importante: Guardar los días
            data.get('almuerzo', True),
            data.get('inicio_almuerzo'), 
            data.get('fin_almuerzo'), 
            almuerzo_min, 
            data.get('tolerancia_ingreso', 0), 
            turno_id
        ))
        conn.commit()
        return {"mensaje": "Turno actualizado correctamente."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar: {str(e)}")
    finally:
        cur.close(); conn.close()

@app.delete("/turnos/{turno_id}")
async def eliminar_turno(turno_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        # Primero, verificamos si hay empleados usando este turno (Protección adicional)
        cur.execute(f"SELECT id FROM {schema}.empleados WHERE turno_id = %s AND eliminado = FALSE LIMIT 1", (turno_id,))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="No puedes eliminar este turno porque hay empleados activos que lo están usando. Reasígnales otro turno primero.")

        # Si está libre, aplicamos el Soft Delete
        cur.execute(f"UPDATE {schema}.turnos SET eliminado = TRUE WHERE id = %s", (turno_id,))
        conn.commit()
        return {"mensaje": "Turno eliminado exitosamente."}
    except HTTPException:
        # Si fue nuestro error personalizado, lo dejamos pasar tal cual
        conn.rollback()
        raise
    except psycopg2.IntegrityError:
         conn.rollback()
         raise HTTPException(status_code=409, detail="Error de Integridad: Este turno está vinculado a registros históricos.")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 10. MÓDULO: VACACIONES Y PERMISOS
# ==============================================================================

# ⚡ NUEVO: Estadísticas Generales de Ausencias (Milisegundos)
# ⚡ NUEVO: Estadísticas Generales y Cargos (Milisegundos)
@app.get("/ausencias/stats")
async def obtener_ausencias_stats(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # 1. Contadores
        cur.execute(f"""
            SELECT 
                COUNT(e.id) as todos,
                COUNT(a.id) FILTER (WHERE a.tipo = 'vacacion') as vacaciones,
                COUNT(a.id) FILTER (WHERE a.tipo = 'permiso') as permisos
            FROM {schema}.empleados e
            LEFT JOIN {schema}.ausencias a ON a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin
            WHERE e.eliminado = FALSE AND e.activo = TRUE
        """)
        res = cur.fetchone()
        
        # 2. Cargos Únicos
        cur.execute(f"SELECT DISTINCT cargo FROM {schema}.empleados WHERE eliminado = FALSE AND activo = TRUE AND cargo IS NOT NULL AND cargo != '' ORDER BY cargo")
        cargos = [row['cargo'] for row in cur.fetchall()]

        todos = res['todos'] or 0
        vacaciones = res['vacaciones'] or 0
        permisos = res['permisos'] or 0
        trabajando = todos - vacaciones - permisos
        
        return {
            "todos": todos, "vacaciones": vacaciones, "permisos": permisos, "trabajando": trabajando,
            "cargos": cargos # ⚡ Ahora el backend entrega la lista limpia
        }
    finally:
        cur.close()
        conn.close()

# ⚡ NUEVO: Buscador de Empleados con Cálculo Kardex "On-The-Fly"
@app.get("/ausencias/directorio")
async def buscar_directorio_ausencias(
    estado: str = "todos", q: str = "", sucursal_id: str = "", seccion_id: str = "", cargo: str = "", limite: int = 100, usuario = Depends(verificar_token)
):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # 1. Traemos la lista con la pre-suma de vacaciones (Todo en 1 sola consulta)
        query = f"""
            SELECT e.id, e.bio_id, e.nombres, e.apellidos, e.ci, e.cargo, e.tipo_contrato,
                   s.nombre as sucursal_nombre, sec.nombre as seccion_nombre,
                   e.fecha_antiguedad, e.saldo_vacaciones_inicial,
                   (SELECT tipo FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin LIMIT 1) as estado_ausencia,
                   COALESCE((SELECT SUM(dias_descontados) FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.tipo = 'vacacion' AND a.estado = 'aprobado' AND a.eliminado = FALSE), 0) as dias_tomados
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            WHERE e.eliminado = FALSE AND e.activo = TRUE
        """
        parametros = []
        
        if estado == "vacaciones": query += f" AND (SELECT tipo FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin LIMIT 1) = 'vacacion'"
        elif estado == "permisos": query += f" AND (SELECT tipo FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin LIMIT 1) = 'permiso'"
        elif estado == "trabajando": query += f" AND (SELECT tipo FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND CURRENT_DATE BETWEEN a.fecha_inicio AND a.fecha_fin LIMIT 1) IS NULL"
        
        if sucursal_id and sucursal_id.isdigit(): query += " AND e.sucursal_id = %s"; parametros.append(int(sucursal_id))
        if seccion_id and seccion_id.isdigit(): query += " AND e.seccion_id = %s"; parametros.append(int(seccion_id))
        if cargo: query += " AND e.cargo = %s"; parametros.append(cargo)
        if q:
            query += " AND (e.nombres ILIKE %s OR e.apellidos ILIKE %s OR e.ci ILIKE %s OR CAST(e.bio_id AS TEXT) ILIKE %s)"
            termino = f"%{q}%"
            parametros.extend([termino, termino, termino, termino])
            
        query += " ORDER BY e.nombres ASC LIMIT %s"
        parametros.append(limite)
        cur.execute(query, tuple(parametros))
        empleados = cur.fetchall()

        # 2. Motor Python en Ram: Calculamos los días disponibles para cada uno sin saturar la DB
        hoy = date.today()
        for emp in empleados:
            emp["dias_disponibles_calculado"] = "0.00"
            if emp["fecha_antiguedad"]:
                dif = relativedelta(hoy, emp["fecha_antiguedad"])
                anios, meses = dif.years, dif.months
                dpa = 15 if anios < 5 else (20 if anios < 10 else 30)
                acumulado = float(emp["saldo_vacaciones_inicial"] or 0) + (anios * dpa) + ((dpa / 12) * meses)
                disp = round(acumulado - float(emp["dias_tomados"]), 2)
                emp["dias_disponibles_calculado"] = f"{disp:.2f}"
            else:
                emp["dias_disponibles_calculado"] = "Sin Fecha"

            if emp["fecha_antiguedad"]: emp["fecha_antiguedad"] = str(emp["fecha_antiguedad"])
            
        return empleados
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 1. KARDEX (Calcula los días disponibles matemáticamente)
# ------------------------------------------------------------------------------
@app.get("/empleados/{empleado_id}/kardex_vacaciones")
async def calcular_vacaciones(empleado_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Traer datos del empleado
        cur.execute(f"SELECT fecha_antiguedad, saldo_vacaciones_inicial FROM {schema}.empleados WHERE id = %s", (empleado_id,))
        emp = cur.fetchone()
        
        if not emp or not emp["fecha_antiguedad"]:
            return {"dias_disponibles": 0, "mensaje": "Sin fecha de antigüedad configurada"}

        fecha_ant = emp["fecha_antiguedad"]
        saldo_inicial = float(emp["saldo_vacaciones_inicial"] or 0)
        hoy = date.today()

        # Calcular tiempo de servicio
        diferencia = relativedelta(hoy, fecha_ant)
        anios_totales = diferencia.years
        meses_extra = diferencia.months

        # Escala Laboral
        if anios_totales < 5:
            dias_por_anio = 15
        elif anios_totales < 10:
            dias_por_anio = 20
        else:
            dias_por_anio = 30

        # Cálculo Progresivo (Devengo mensual)
        dias_ganados_por_anios = anios_totales * dias_por_anio
        dias_ganados_por_meses = (dias_por_anio / 12) * meses_extra
        total_acumulado = saldo_inicial + dias_ganados_por_anios + dias_ganados_por_meses

        # Restar los días ya tomados
        cur.execute(f"""
            SELECT SUM(dias_descontados) as tomados 
            FROM {schema}.ausencias 
            WHERE empleado_id = %s AND tipo = 'vacacion' AND estado = 'aprobado' AND eliminado = FALSE
        """, (empleado_id,))
        tomados_db = cur.fetchone()
        total_tomados = float(tomados_db["tomados"] or 0)

        dias_disponibles = round(total_acumulado - total_tomados, 2)

        return {
            "antiguedad": f"{anios_totales} años y {meses_extra} meses",
            "tasa_actual": f"{dias_por_anio} días/año",
            "total_acumulado": round(total_acumulado, 2),
            "total_tomados": total_tomados,
            "dias_disponibles": dias_disponibles
        }
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 2. HISTORIAL DE AUSENCIAS (Trae la lista de vacaciones/permisos tomados)
# ------------------------------------------------------------------------------
@app.get("/empleados/{empleado_id}/ausencias")
async def obtener_historial_ausencias(empleado_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # ⚡ FIX: Agregamos 'requiere_reposicion' a la consulta para que el Calendario despierte el KPI de deudas
        cur.execute(f"""
            SELECT id, tipo, fecha_inicio, fecha_fin, hora_inicio, hora_fin, 
                   horas_totales, dias_descontados, motivo, estado, requiere_reposicion
            FROM {schema}.ausencias
            WHERE empleado_id = %s AND eliminado = FALSE
            ORDER BY fecha_inicio DESC
        """, (empleado_id,))
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 3. SEGURIDAD: AJUSTAR SALDO INICIAL (Requiere Contraseña de Admin)
# ------------------------------------------------------------------------------
# ⚡ ESTA ES LA NUEVA RUTA QUE FALTABA PARA QUE FUNCIONE EL BOTÓN DEL FRONTEND
@app.put("/empleados/{empleado_id}/saldo_inicial")
async def actualizar_saldo_inicial(empleado_id: int, data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    nuevo_saldo = data.get("saldo_inicial")
    admin_password = data.get("admin_password")

    if nuevo_saldo is None or not admin_password:
        raise HTTPException(status_code=409, detail="El saldo y la contraseña son obligatorios.")

    # Verificación de Seguridad (Contraseña en Base Maestra)
    conn_pub = conectar_bd("public")
    cur_pub = conn_pub.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur_pub.execute("SELECT password_hash FROM usuarios WHERE id = %s", (usuario["id"],))
    user_db = cur_pub.fetchone()
    cur_pub.close()
    conn_pub.close()

    # Comparamos la contraseña digitada
    if not user_db or not bcrypt.checkpw(admin_password.encode(), user_db["password_hash"].encode()):
        raise HTTPException(status_code=409, detail="Contraseña de Administrador incorrecta. Operación denegada.")

    # Guardar el nuevo saldo
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        cur.execute(f"""
            UPDATE {schema}.empleados 
            SET saldo_vacaciones_inicial = %s 
            WHERE id = %s
        """, (nuevo_saldo, empleado_id))
        conn.commit()
        return {"mensaje": "Saldo inicial ajustado correctamente."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 4. REGISTRAR UNA NUEVA AUSENCIA (Motor de Guardado Inteligente)
# ------------------------------------------------------------------------------
@app.post("/ausencias")
async def registrar_ausencia(data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) 
    try:
        empleado_id = data.get("empleado_id")
        tipo = data.get("tipo")  
        fecha_inicio = data.get("fecha_inicio")
        fecha_fin = data.get("fecha_fin")
        motivo = data.get("motivo", "")
        
        # ⚡ Nuevos campos recibidos desde el Frontend
        por_dias = data.get("por_dias", False)
        requiere_reposicion = data.get("requiere_reposicion", False)

        dias_descontados = 0
        horas_totales = 0
        hora_inicio = None
        hora_fin = None

        # ⚡ LÓGICA MEJORADA: Si es vacación O permiso por días completos
        # ⚡ LÓGICA MEJORADA: Si es vacación O permiso por días completos
        if tipo == "vacacion" or (tipo == "permiso" and por_dias):
            f_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            f_fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
            
            cur.execute(f"SELECT * FROM {schema}.empleados e JOIN {schema}.turnos t ON e.turno_id = t.id WHERE e.id = %s", (empleado_id,))
            turno_emp = cur.fetchone()

            if not turno_emp:
                raise HTTPException(status_code=409, detail="El empleado no tiene un turno asignado.")

            dias_laborales_json = turno_emp["dias"] 
            es_medio_tiempo_fines = turno_emp["medio_tiempo_fines"]
            mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}
            dia_actual = f_inicio
            
            # ⚡ Variables para el motor matemático del ERP
            minutos_totales_deuda = 0
            
            while dia_actual <= f_fin:
                letra_dia = mapa_dias[dia_actual.weekday()]
                if dias_laborales_json.get(letra_dia, False):
                    # 1. Contamos el día para el Frontend y Kardex
                    factor_dia = 0.5 if (letra_dia in ['S', 'D'] and es_medio_tiempo_fines) else 1.0
                    dias_descontados += factor_dia
                    
                    # 2. ⚡ CALCULAMOS LAS HORAS EXACTAS DE ESE DÍA PARA LA DEUDA
                    # Convertimos todo a minutos para evitar errores de reloj
                    in_h, in_m = turno_emp["hora_ingreso"].hour, turno_emp["hora_ingreso"].minute
                    out_h, out_m = turno_emp["hora_salida"].hour, turno_emp["hora_salida"].minute
                    
                    mins_ingreso = (in_h * 60) + in_m
                    mins_salida = (out_h * 60) + out_m
                    if mins_salida < mins_ingreso: mins_salida += 1440 # Cruce de madrugada
                    
                    mins_trabajo_dia = mins_salida - mins_ingreso
                    
                    # Restamos el almuerzo de este día específico
                    if turno_emp["almuerzo"] and turno_emp.get("hora_inicio_almuerzo"):
                        alm_in_h, alm_in_m = turno_emp["hora_inicio_almuerzo"].hour, turno_emp["hora_inicio_almuerzo"].minute
                        alm_out_h, alm_out_m = turno_emp["hora_fin_almuerzo"].hour, turno_emp["hora_fin_almuerzo"].minute
                        
                        mins_alm_in = (alm_in_h * 60) + alm_in_m
                        mins_alm_out = (alm_out_h * 60) + alm_out_m
                        if mins_alm_out < mins_alm_in: mins_alm_out += 1440
                        
                        mins_trabajo_dia -= (mins_alm_out - mins_alm_in)
                    
                    # Aplicamos factor de medio tiempo si es fin de semana
                    mins_trabajo_dia = mins_trabajo_dia * factor_dia
                    
                    # Sumamos a la bolsa total de deuda
                    minutos_totales_deuda += mins_trabajo_dia

                dia_actual += timedelta(days=1) 

            if dias_descontados == 0:
                raise HTTPException(status_code=409, detail="El rango no contiene días laborales.")
                
            # ⚡ Guardamos las horas exactas en la BD
            horas_totales = round(minutos_totales_deuda / 60.0, 2)
        else:
            # ⚡ MOTOR ERP: Lógica de Permiso por Horas Multidía y Almuerzos
            fecha_ini_str = data.get("fecha_inicio_permiso")
            hora_ini_str = data.get("hora_inicio_permiso")
            fecha_fin_str = data.get("fecha_fin_permiso")
            hora_fin_str = data.get("hora_fin_permiso")
            
            # Unificamos Fechas y Horas para cálculo
            dt_inicio = datetime.strptime(f"{fecha_ini_str} {hora_ini_str}", "%Y-%m-%d %H:%M")
            dt_fin = datetime.strptime(f"{fecha_fin_str} {hora_fin_str}", "%Y-%m-%d %H:%M")
            
            if dt_fin < dt_inicio:
                raise HTTPException(status_code=409, detail="La fecha de retorno no puede ser en el pasado.")

            cur.execute(f"SELECT t.* FROM {schema}.empleados e JOIN {schema}.turnos t ON e.turno_id = t.id WHERE e.id = %s", (empleado_id,))
            turno_emp = cur.fetchone()

            segundos_totales = 0
            mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}
            dias_json = turno_emp["dias"]
            
            def combinar_dt(d, t): return datetime.combine(d, t) if t else None

            dia_cursor = dt_inicio.date()
            while dia_cursor <= dt_fin.date():
                letra_dia = mapa_dias[dia_cursor.weekday()]
                
                # Si ese día el empleado trabaja, calculamos cruces
                if dias_json.get(letra_dia, False):
                    w_start = combinar_dt(dia_cursor, turno_emp["hora_ingreso"])
                    w_end = combinar_dt(dia_cursor, turno_emp["hora_salida"])
                    
                    if w_end < w_start: w_end += timedelta(days=1) # Turnos nocturnos
                    
                    # Intersección entre [El permiso] y [El turno laboral]
                    p_start = max(w_start, dt_inicio)
                    p_end = min(w_end, dt_fin)
                    
                    if p_start < p_end:
                        overlap_secs = (p_end - p_start).total_seconds()
                        
                        # ⚡ RESTA DEL ALMUERZO SI SE CRUZA
                        if turno_emp["almuerzo"] and turno_emp.get("hora_inicio_almuerzo"):
                            l_start = combinar_dt(dia_cursor, turno_emp["hora_inicio_almuerzo"])
                            l_end = combinar_dt(dia_cursor, turno_emp["hora_fin_almuerzo"])
                            if l_end < l_start: l_end += timedelta(days=1)
                            
                            lunch_o_start = max(p_start, l_start)
                            lunch_o_end = min(p_end, l_end)
                            if lunch_o_start < lunch_o_end:
                                overlap_secs -= (lunch_o_end - lunch_o_start).total_seconds()
                                
                        segundos_totales += overlap_secs
                dia_cursor += timedelta(days=1)
            
            horas_totales = round(segundos_totales / 3600.0, 2)
            
            # ⚡ NUEVA BARRERA DE SEGURIDAD ERP
            if horas_totales <= 0:
                raise HTTPException(
                    status_code=409, 
                    detail="Operación denegada: El permiso cae en el día libre del empleado, está fuera de su horario de trabajo, o es anulado por su hora de almuerzo. Horas a descontar: 0."
                )

            # Formateamos para guardar en DB
            fecha_inicio = fecha_ini_str
            fecha_fin = fecha_fin_str
            hora_inicio = hora_ini_str
            hora_fin = hora_fin_str

        # ⚡ INSERT ACTUALIZADO (Incluye requiere_reposicion)
        cur.execute(f"""
            INSERT INTO {schema}.ausencias 
            (empleado_id, tipo, fecha_inicio, fecha_fin, hora_inicio, hora_fin, horas_totales, dias_descontados, motivo, requiere_reposicion)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (empleado_id, tipo, fecha_inicio, fecha_fin, hora_inicio, hora_fin, horas_totales, dias_descontados, motivo, requiere_reposicion))
        
        conn.commit()

        # ⚡ DISPARADOR: Recalculamos cada día afectado por la ausencia
        dia_loop = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        
        while dia_loop <= fecha_fin_dt:
            procesar_asistencia_dia(schema, empleado_id, dia_loop)
            dia_loop += timedelta(days=1)

        return {"mensaje": f"{tipo.capitalize()} registrada correctamente."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 5. ANULAR AUSENCIA (Soft Delete con Protección de Auditoría)
# ------------------------------------------------------------------------------
@app.delete("/ausencias/{ausencia_id}")
async def anular_ausencia(ausencia_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(f"SELECT fecha_inicio FROM {schema}.ausencias WHERE id = %s", (ausencia_id,))
        ausencia = cur.fetchone()
        
        if not ausencia:
            raise HTTPException(status_code=404, detail="Registro no encontrado.")
            
        hoy = date.today()
        if ausencia["fecha_inicio"] <= hoy and usuario["rol"] != "superadmin":
             raise HTTPException(status_code=403, detail="No puedes borrar ausencias pasadas o en curso. Solo el SuperAdministrador tiene este privilegio.")

        cur.execute(f"UPDATE {schema}.ausencias SET eliminado = TRUE, estado = 'anulado' WHERE id = %s", (ausencia_id,))
        
        conn.commit()
        return {"mensaje": "Registro anulado. El saldo ha sido restituido al empleado."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ------------------------------------------------------------------------------
# 6. EDITAR OBSERVACIONES DE AUSENCIA
# ------------------------------------------------------------------------------
@app.put("/ausencias/{ausencia_id}")
async def editar_ausencia(ausencia_id: int, data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        nuevo_motivo = data.get("motivo")
        
        cur.execute(f"""
            UPDATE {schema}.ausencias 
            SET motivo = %s 
            WHERE id = %s
        """, (nuevo_motivo, ausencia_id))
        
        conn.commit()
        return {"mensaje": "Observaciones actualizadas."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=422, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ⚡ NUEVO: FABRICANTE DE BOLETA DE PERMISO/VACACIÓN (Server-Side PDF)
@app.get("/ausencias/{ausencia_id}/boleta/pdf")
async def descargar_boleta_pdf(ausencia_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(f"""
            SELECT a.*, e.nombres, e.apellidos, e.ci, e.bio_id, e.cargo
            FROM {schema}.ausencias a
            JOIN {schema}.empleados e ON a.empleado_id = e.id
            WHERE a.id = %s
        """, (ausencia_id,))
        reg = cur.fetchone()
        if not reg: raise HTTPException(status_code=404, detail="Registro no encontrado")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=portrait(A5), leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()

        # 1. Cabecera (Azul)
        elements.append(Paragraph(f"<font color='#1e3a8a' size='14'><b>BOLETA DE AUTORIZACIÓN</b></font>", styles['Title']))
        elements.append(Paragraph("<br/>", styles['Normal']))

        # 2. Datos Empleado
        elements.append(Paragraph(f"<b>Empleado:</b> {reg['nombres']} {reg['apellidos']}", styles['Normal']))
        elements.append(Paragraph(f"<b>C.I.:</b> {reg['ci']} &nbsp;&nbsp;&nbsp; <b>ID Reloj:</b> {reg.get('bio_id') or 'N/A'}", styles['Normal']))
        elements.append(Paragraph(f"<b>Cargo:</b> {reg.get('cargo') or 'No especificado'}", styles['Normal']))
        elements.append(Paragraph("<br/><hr/><br/>", styles['Normal']))

        # 3. Detalle Solicitud
        elements.append(Paragraph(f"<b>Detalle de la Solicitud:</b>", styles['Normal']))
        elements.append(Paragraph(f"<b>Tipo:</b> {reg['tipo'].upper()}", styles['Normal']))
        elements.append(Paragraph(f"<b>Fechas:</b> {reg['fecha_inicio']} al {reg['fecha_fin']}", styles['Normal']))

        if reg['tipo'] == 'permiso':
            if reg.get('hora_inicio'):
                elements.append(Paragraph(f"<b>Horario:</b> {str(reg['hora_inicio'])[:5]} a {str(reg['hora_fin'])[:5]}", styles['Normal']))
            else:
                elements.append(Paragraph(f"<b>Horario:</b> Jornada Completa", styles['Normal']))
            
            txt_horas = f"<b>Total Horas:</b> {reg['horas_totales']} hrs"
            if reg.get('requiere_reposicion'):
                txt_horas += " &nbsp;&nbsp;&nbsp; <font color='red'><b>* DEBE REPONER HORAS *</b></font>"
            elements.append(Paragraph(txt_horas, styles['Normal']))
        else:
            elements.append(Paragraph(f"<b>Total Días Descontados:</b> {reg['dias_descontados']}", styles['Normal']))

        # 4. Observaciones
        elements.append(Paragraph("<br/><b>Observaciones:</b>", styles['Normal']))
        obs = reg.get('motivo') or "Sin observaciones registradas."
        elements.append(Paragraph(f"<i>{obs}</i>", styles['Italic']))

        # 5. Firmas
        elements.append(Paragraph("<br/><br/><br/><br/>", styles['Normal']))
        firmas_data = [[Paragraph("_________________________<br/>Firma del Empleado", styles['Normal']),
                        Paragraph("_________________________<br/>Autorizado por (RRHH)", styles['Normal'])]]
        t = Table(firmas_data, colWidths=[150, 150])
        elements.append(t)
        
        doc.build(elements)
        output.seek(0)
        return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=Boleta_{reg['ci']}.pdf"})
    finally:
        cur.close()
        conn.close()

# ⚡ NUEVO: FABRICANTE DE KARDEX DE HISTORIAL (Server-Side PDF)
@app.get("/empleados/{empleado_id}/historial_ausencias/pdf")
async def descargar_historial_ausencias_pdf(empleado_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute(f"SELECT * FROM {schema}.empleados WHERE id = %s", (empleado_id,))
        emp = cur.fetchone()
        
        cur.execute(f"""
            SELECT * FROM {schema}.ausencias 
            WHERE empleado_id = %s AND eliminado = FALSE 
            ORDER BY fecha_inicio DESC
        """, (empleado_id,))
        historial = cur.fetchall()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph(f"<b>REPORTE KARDEX DE AUSENCIAS</b>", styles['Title']))
        elements.append(Paragraph(f"<b>Empleado:</b> {emp['nombres']} {emp['apellidos']} ({emp['ci']})", styles['Normal']))
        elements.append(Paragraph(f"<b>Cargo:</b> {emp.get('cargo') or 'N/A'} <br/><br/>", styles['Normal']))

        if not historial:
            elements.append(Paragraph("El empleado no tiene registros de ausencias.", styles['Normal']))
        else:
            data = [["TIPO", "FECHA / HORARIO", "DESCUENTO", "ESTADO", "OBSERVACIONES"]]
            for reg in historial:
                tiempo = f"{reg['dias_descontados']} Días" if reg['tipo'] == 'vacacion' else f"{reg['horas_totales']} Hrs"
                fechas = f"{reg['fecha_inicio']}"
                if reg['fecha_inicio'] != reg['fecha_fin']: fechas += f" al {reg['fecha_fin']}"
                
                if reg['tipo'] == 'permiso':
                    if reg.get('hora_inicio'): fechas += f"\n({str(reg['hora_inicio'])[:5]} a {str(reg['hora_fin'])[:5]})"
                    else: fechas += "\n(Día Completo)"

                estado_txt = reg['estado'].upper()
                if reg['tipo'] == 'permiso' and reg.get('requiere_reposicion'): estado_txt += "\n(REPONER HORAS)"

                data.append([
                    reg['tipo'].upper(), fechas, tiempo, estado_txt, reg.get('motivo') or 'Ninguna'
                ])

            t = Table(data, colWidths=[60, 150, 70, 90, 150])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))
            elements.append(t)

        doc.build(elements)
        output.seek(0)
        return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=Kardex_{emp['ci']}.pdf"})
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 11. MOTOR MATEMÁTICO DE ASISTENCIA (CORE ERP)
# ==============================================================================

def calcular_dia_asistencia(marcajes_brutos: list, turno: dict, permisos: list, salario_base: float, fecha_dia: date = None):
    marcajes_limpios = []
    for m in sorted(marcajes_brutos):
        if not marcajes_limpios or (m - marcajes_limpios[-1]).total_seconds() > 180:
            marcajes_limpios.append(m)

    if not fecha_dia: fecha_dia = date.today()

    resumen = {
        "hora_entrada": marcajes_limpios[0].time() if len(marcajes_limpios) > 0 else None,
        "hora_inicio_almuerzo": marcajes_limpios[1].time() if len(marcajes_limpios) > 1 else None,
        "hora_fin_almuerzo": marcajes_limpios[2].time() if len(marcajes_limpios) > 2 else None,
        # ⚡ FIX: Solo asignamos Salida si hay más de 1 marcaje, evitando el clonado 09:20 / 09:20
        "hora_salida": marcajes_limpios[-1].time() if len(marcajes_limpios) > 1 else None,
        "minutos_retraso_entrada": 0,
        "minutos_exceso_almuerzo": 0,
        "estado": "Falta",
        "deuda_generada_bs": 0.00,
        "horas_trabajadas": 0.00,
        "horas_permiso_dia": 0.00,
        "horas_extras": 0.00
    }

    # ⚡ FIX 1: Identificador universal de día libre
    mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}
    letra_dia = mapa_dias[fecha_dia.weekday()]
    dias_json = turno.get('dias', {})
    if isinstance(dias_json, str): 
        import json
        try: dias_json = json.loads(dias_json)
        except: dias_json = {}
    es_dia_laboral = dias_json.get(letra_dia, False)

    # 1. Establecer el Bloque del Turno Total
    t_in = datetime.combine(fecha_dia, turno['hora_ingreso'])
    t_out = datetime.combine(fecha_dia, turno['hora_salida'])
    if t_out < t_in: t_out += timedelta(days=1)

    # 2. Establecer el Muro del Almuerzo
    tiene_almuerzo = turno.get('almuerzo', False)
    if tiene_almuerzo and turno.get('hora_inicio_almuerzo') and turno.get('hora_fin_almuerzo'):
        l_in = datetime.combine(fecha_dia, turno['hora_inicio_almuerzo'])
        l_out = datetime.combine(fecha_dia, turno['hora_fin_almuerzo'])
        if l_in < t_in: l_in += timedelta(days=1)
        if l_out < l_in: l_out += timedelta(days=1)
    else:
        l_in = t_out
        l_out = t_out # Si no hay almuerzo, colapsamos el muro al final del turno

    nueva_hora_entrada_oficial = t_in
    min_cubiertos_permiso = 0
    p_in = t_out
    p_out = t_out

    # 3. Mapeo del Bloque de Permiso
    if permisos:
        for p in permisos:
            h_ini = p.get('hora_inicio') if p.get('hora_inicio') is not None else time(0, 0)
            h_fin = p.get('hora_fin') if p.get('hora_fin') is not None else time(23, 59)

            f_ini_raw = p.get('fecha_inicio', fecha_dia)
            f_ini = f_ini_raw.date() if hasattr(f_ini_raw, 'date') and callable(getattr(f_ini_raw, 'date')) else (date.fromisoformat(f_ini_raw[:10]) if isinstance(f_ini_raw, str) else f_ini_raw)
            f_fin_raw = p.get('fecha_fin', fecha_dia)
            f_fin = f_fin_raw.date() if hasattr(f_fin_raw, 'date') and callable(getattr(f_fin_raw, 'date')) else (date.fromisoformat(f_fin_raw[:10]) if isinstance(f_fin_raw, str) else f_fin_raw)

            p_in = max(t_in, datetime.combine(f_ini, h_ini))
            p_out = min(t_out, datetime.combine(f_fin, h_fin))
            
            if p_in < p_out:
                min_cubiertos_permiso += (p_out - p_in).total_seconds() / 60
                if p_in <= t_in <= p_out:
                    nueva_hora_entrada_oficial = p_out

    # ⚡ FIX: EL SALTO DEL MURO DE ALMUERZO
    # Si el permiso empujó la hora oficial de entrada de modo que caiga DENTRO del horario de almuerzo,
    # el sistema perdona ese tiempo y espera al empleado directamente al finalizar el almuerzo.
    if tiene_almuerzo and l_in <= nueva_hora_entrada_oficial < l_out:
        nueva_hora_entrada_oficial = l_out

    resumen["horas_permiso_dia"] = round(min_cubiertos_permiso / 60, 2)
    es_permiso_total = min_cubiertos_permiso >= ((t_out - t_in).total_seconds() / 60 * 0.9)

    # ⚡ FIX 3: EL MOTOR TOPOLÓGICO PARA MARCAS ESPERADAS
    def minutos_libres(b_start, b_end):
        """Calcula los minutos libres en un bloque, restando la sombra del permiso"""
        if b_end <= b_start: return 0
        overlap_start = max(b_start, p_in)
        overlap_end = min(b_end, p_out)
        overlap_mins = max(0, (overlap_end - overlap_start).total_seconds() / 60)
        return max(0, ((b_end - b_start).total_seconds() / 60) - overlap_mins)

    min_libres_manana = minutos_libres(t_in, l_in)
    min_libres_tarde = minutos_libres(l_out, t_out)

    if es_permiso_total:
        marcajes_esperados = 0
    elif min_libres_manana > 0 and min_libres_tarde > 0:
        marcajes_esperados = 4 if tiene_almuerzo else 2
    else:
        # Si el permiso se comió la mañana o la tarde entera, solo queda un bloque de trabajo
        marcajes_esperados = 2

    # Veredicto si no hay marcas
    if not marcajes_limpios:
        if es_permiso_total:
            resumen["estado"] = "Permiso"
        elif not es_dia_laboral:
            resumen["estado"] = "Descanso" # ⚡ FIX 1 APLICADO
        elif t_out > datetime.now():
            resumen["estado"] = "Pendiente"
        else:
            resumen["estado"] = "Falta"
        return resumen

    # Cálculos normales
    if resumen["hora_entrada"]:
        dt_entrada_real = datetime.combine(fecha_dia, resumen["hora_entrada"])
        retraso_seg = (dt_entrada_real - nueva_hora_entrada_oficial).total_seconds()
        if retraso_seg > (turno.get('tolerancia_min', 0) * 60):
            resumen["minutos_retraso_entrada"] = int(retraso_seg / 60)

    # 7. Cálculo de Horas Trabajadas Reales
    if len(marcajes_limpios) >= 2:
        dt_first = marcajes_limpios[0]
        dt_last = marcajes_limpios[-1]
        segundos_brutos = (dt_last - dt_first).total_seconds()
        
        # ⚡ FIX 1: Descuento real del almuerzo (Destruyendo las 11 horas)
        if marcajes_esperados == 4 and len(marcajes_limpios) >= 3:
            if resumen["hora_inicio_almuerzo"] and resumen["hora_fin_almuerzo"]:
                dt_alm_in = datetime.combine(fecha_dia, resumen["hora_inicio_almuerzo"])
                dt_alm_out = datetime.combine(fecha_dia, resumen["hora_fin_almuerzo"])
                if dt_alm_out < dt_alm_in: dt_alm_out += timedelta(days=1)
                
                # Cuánto tiempo estuvo comiendo realmente
                segundos_almuerzo_real = max(0, (dt_alm_out - dt_alm_in).total_seconds())
                segundos_brutos -= segundos_almuerzo_real
        
        resumen["horas_trabajadas"] = round(max(0, segundos_brutos) / 3600.0, 2)

    # ⚡ FIX: BOLSILLO "C" (Tiempo Extra Real - Fuera de Fronteras)
    # Comparamos las huellas reales (datetime) contra los límites del turno (t_in y t_out)
    minutos_extra_total = 0
    if len(marcajes_limpios) > 0:
        dt_in_real = marcajes_limpios[0]
        if dt_in_real < t_in: # Llegó antes de su turno
            minutos_extra_total += (t_in - dt_in_real).total_seconds() / 60
            
    if len(marcajes_limpios) >= 2:
        dt_out_real = marcajes_limpios[-1]
        if dt_out_real > t_out: # Se quedó después de su turno
            minutos_extra_total += (dt_out_real - t_out).total_seconds() / 60

    resumen["horas_extras"] = round(minutos_extra_total / 60, 2)

    # ⚡ FIX: Veredicto de Estado Inteligente (Eliminamos el bloque duplicado)
    conteo = len(marcajes_limpios)
    if conteo < marcajes_esperados:
        # Si le faltan marcas, pero su turno aún no termina, está "Trabajando"
        if t_out > datetime.now():
            resumen["estado"] = "Trabajando"
        else:
            resumen["estado"] = "Incompleto"
    else:
        resumen["estado"] = "Tarde" if resumen["minutos_retraso_entrada"] > 0 else "Puntual"

    if turno.get('descuento', True):
        valor_minuto = (float(salario_base) / 30 / 8 / 60) if salario_base > 0 else 0
        resumen["deuda_generada_bs"] = round(resumen["minutos_retraso_entrada"] * valor_minuto, 2)

    return resumen

# ==============================================================================
# 12. LECTURA DE ASISTENCIA (PARA EL CALENDARIO FRONTEND)
# ==============================================================================

@app.get("/empleados/{empleado_id}/asistencia/{anio}/{mes}")
async def obtener_asistencia_mensual(empleado_id: int, anio: int, mes: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        import calendar
        from datetime import date, timedelta

        # 1. LIMPIEZA PEREZOSA (Revisa si hay días trancados)
        cur.execute(f"SELECT fecha, estado, horas_trabajadas FROM {schema}.asistencia_diaria WHERE empleado_id = %s AND EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s", (empleado_id, anio, mes))
        dias_check = cur.fetchall()
        hoy_srv = date.today()
        
        for dc in dias_check:
            congelado = dc["estado"] in ["Trabajando", "Pendiente"] and dc["fecha"] < hoy_srv
            viejo_sin_horas = dc["estado"] not in ["Falta", "Pendiente"] and float(dc["horas_trabajadas"] or 0) == 0.0
            if congelado or viejo_sin_horas:
                procesar_asistencia_dia(schema, empleado_id, dc["fecha"])

        # 2. CARGAR DATA: Note el "t.nombre as turno_nombre"
        cur.execute(f"""
            SELECT e.fecha_ingreso, s.ciudad as sucursal_ciudad, t.nombre as turno_nombre, t.* FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.id = %s
        """, (empleado_id,))
        emp_data = cur.fetchone()

        cur.execute(f"SELECT * FROM {schema}.asistencia_diaria WHERE empleado_id = %s AND EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s", (empleado_id, anio, mes))
        asistencia_raw = {str(d["fecha"]): d for d in cur.fetchall()}

        _, dias_del_mes = calendar.monthrange(anio, mes)
        cur.execute(f"""
            SELECT * FROM {schema}.ausencias 
            WHERE empleado_id = %s AND eliminado = FALSE AND estado = 'aprobado'
            AND ((EXTRACT(YEAR FROM fecha_inicio) = %s AND EXTRACT(MONTH FROM fecha_inicio) = %s)
                 OR (EXTRACT(YEAR FROM fecha_fin) = %s AND EXTRACT(MONTH FROM fecha_fin) = %s)
                 OR (fecha_inicio <= %s AND fecha_fin >= %s))
        """, (empleado_id, anio, mes, anio, mes, f"{anio}-{mes:02d}-{dias_del_mes:02d}", f"{anio}-{mes:02d}-01"))
        ausencias_raw = cur.fetchall()

        cur.execute(f"SELECT * FROM {schema}.feriados WHERE eliminado = FALSE")
        feriados_dict = {}
        for f in cur.fetchall():
            f_date = str(f['fecha'])
            f_md = f_date[5:]
            if f.get('recurrente'): feriados_dict[f_md] = f
            else: feriados_dict[f_date] = f

        # 3. 🧠 MOTOR MATEMÁTICO
        primer_dia_semana = date(anio, mes, 1).weekday()
        vacios_inicio = primer_dia_semana
        
        horas_turno_base = 0.0
        paga_extras = False
        dias_laborales_json = {}
        es_medio_tiempo_fines = False
        
        if emp_data and emp_data.get('hora_ingreso') and emp_data.get('hora_salida'):
            import json
            raw_dias = emp_data.get('dias', '{}')
            try: dias_laborales_json = json.loads(raw_dias) if isinstance(raw_dias, str) else raw_dias
            except: dias_laborales_json = {}
            
            paga_extras = emp_data.get('horas_extras', False)
            es_medio_tiempo_fines = emp_data.get('medio_tiempo_fines', False)
            
            in_h, in_m = emp_data['hora_ingreso'].hour, emp_data['hora_ingreso'].minute
            out_h, out_m = emp_data['hora_salida'].hour, emp_data['hora_salida'].minute
            min_in = in_h * 60 + in_m
            min_out = out_h * 60 + out_m
            if min_out < min_in: min_out += 24 * 60
            mins = min_out - min_in
            if emp_data.get('almuerzo'): mins -= emp_data.get('almuerzo_min', 0)
            horas_turno_base = round(mins / 60.0, 2)

        sum_horas_esperadas = sum_horas_trabajadas = sum_horas_extras = sum_horas_reponer = 0.0
        dias_trabajados_count = retraso_total_min = 0
        deuda_total_bs = 0.0

        mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}
        meses_cortos = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        dias_calendario = []
        
        for d in range(1, dias_del_mes + 1):
            fecha_loop = date(anio, mes, d)
            fecha_str = f"{anio}-{mes:02d}-{d:02d}"
            mes_dia_str = f"{mes:02d}-{d:02d}"
            letra_dia = mapa_dias[fecha_loop.weekday()]
            
            datos_asistencia = asistencia_raw.get(fecha_str)
            
            obj_feriado = feriados_dict.get(fecha_str) or feriados_dict.get(mes_dia_str)
            nombre_feriado = None
            if obj_feriado:
                alcance = str(obj_feriado.get('tipo', '')).lower()
                ciudad_emp = str(emp_data.get('sucursal_ciudad', '')).lower() if emp_data else ''
                if alcance == 'nacional' or alcance == ciudad_emp: nombre_feriado = obj_feriado['descripcion']

            ausencia_activa = next((a for a in ausencias_raw if str(a['fecha_inicio']) <= fecha_str <= str(a['fecha_fin'])), None)
            es_dia_laboral = bool(dias_laborales_json.get(letra_dia, False)) if dias_laborales_json else (letra_dia not in ['S', 'D'])

            factor_dia = 0.5 if (es_dia_laboral and es_medio_tiempo_fines and letra_dia in ['S', 'D']) else 1.0
            hrs_esperadas_hoy = horas_turno_base * factor_dia

            if es_dia_laboral and fecha_loop <= hoy_srv and not nombre_feriado:
                es_dia_libre_completo = ausencia_activa and (ausencia_activa['tipo'] == 'vacacion' or not ausencia_activa.get('hora_inicio'))
                if not es_dia_libre_completo:
                    hrs_para_sumar = hrs_esperadas_hoy
                    if ausencia_activa and ausencia_activa.get('horas_totales'):
                        hrs_para_sumar = max(0, hrs_para_sumar - float(ausencia_activa['horas_totales']))
                    sum_horas_esperadas += hrs_para_sumar

            h_extras_hoy = 0.0
            if datos_asistencia and float(datos_asistencia.get('horas_trabajadas') or 0) > 0:
                sum_horas_trabajadas += float(datos_asistencia['horas_trabajadas'])
                h_extras_hoy = float(datos_asistencia.get('horas_extras') or 0)
                sum_horas_extras += h_extras_hoy
                dias_trabajados_count += 1
                retraso_total_min += int(datos_asistencia.get('minutos_retraso_entrada', 0))
                deuda_total_bs += float(datos_asistencia.get('deuda_generada_bs', 0))

            # ESTADOS VISUALES
            tipo_dia = "laboral"
            estado_asistencia = "Pendiente"
            permite_clic = True

            # ⚡ Evaluamos primero si el permiso cubre todo el día
            es_dia_libre_completo = ausencia_activa and (ausencia_activa['tipo'] == 'vacacion' or not ausencia_activa.get('hora_inicio'))

            if emp_data and emp_data.get('fecha_ingreso') and fecha_loop < emp_data['fecha_ingreso']:
                tipo_dia = "previo_contrato"
                estado_asistencia = "Previo a Contrato"
                permite_clic = False
            elif nombre_feriado and not datos_asistencia:
                tipo_dia = "feriado"
                estado_asistencia = nombre_feriado
                permite_clic = False
            elif es_dia_libre_completo:
                # ⚡ FIX: Prioridad visual absoluta a las ausencias de día completo
                tipo_dia = ausencia_activa['tipo']
                estado_asistencia = ausencia_activa['tipo'].capitalize()
            elif datos_asistencia and datos_asistencia.get('estado') != 'Pendiente':
                estado_asistencia = datos_asistencia['estado']
            elif not es_dia_laboral:
                tipo_dia = "descanso"
                estado_asistencia = "Descanso"
            elif fecha_loop < hoy_srv:
                estado_asistencia = "Falta"

            # ⚡ 4. EL BACKEND PREPARA EL UI PERFECTO PARA EL FRONTEND ⚡
            def format_ui_time(h_obj):
                if not h_obj: return "--:--"
                h_str = str(h_obj)[:5]
                
                # Detectamos si es turno nocturno
                if emp_data and emp_data.get('hora_salida') and emp_data.get('hora_ingreso') and emp_data['hora_salida'] < emp_data['hora_ingreso']:
                    from datetime import time
                    
                    # Frontera inteligente: Las 12:00 PM (Mediodía)
                    # Si es menor a las 12:00, es la madrugada del DÍA SIGUIENTE
                    if h_obj < time(12, 0):
                        d_sig = fecha_loop + timedelta(days=1)
                        txt_dia = f"{d_sig.day:02d} {meses_cortos[d_sig.month-1]}."
                        return f'<span class="text-[9px] text-blue-500 font-bold bg-blue-50 px-1 rounded mr-1" title="Día Siguiente">{txt_dia}</span>{h_str}'
                    # Si es mayor a las 12:00, es la tarde/noche del DÍA ACTUAL
                    else:
                        txt_dia = f"{fecha_loop.day:02d} {meses_cortos[fecha_loop.month-1]}."
                        return f'<span class="text-[9px] text-slate-500 font-bold bg-slate-100 border border-slate-200 px-1 rounded mr-1" title="Día Actual">{txt_dia}</span>{h_str}'
                        
                return h_str

            exige_almuerzo = emp_data.get('almuerzo', False) if emp_data else False
            if exige_almuerzo and ausencia_activa and ausencia_activa.get('hora_inicio') and emp_data.get('hora_inicio_almuerzo'):
                # ⚡ FIX: Usamos >= y <= (Inclusivo) para detectar si el permiso "toca" la frontera del almuerzo.
                # Si el permiso termina exactamente a las 12:30 (inicio de almuerzo), lo absorbe.
                if (ausencia_activa['hora_inicio'] <= emp_data['hora_fin_almuerzo']) and (ausencia_activa['hora_fin'] >= emp_data['hora_inicio_almuerzo']):
                    exige_almuerzo = False

            datos_modal = {}
            if permite_clic:
                datos_modal = {
                    "fecha": fecha_str,
                    "estado": estado_asistencia,
                    "horas_turno": hrs_esperadas_hoy,
                    "horas_trabajadas": float(datos_asistencia['horas_trabajadas']) if datos_asistencia else 0.0,
                    "horas_extras": h_extras_hoy,
                    "minutos_retraso_entrada": int(datos_asistencia['minutos_retraso_entrada']) if datos_asistencia else 0,
                    "deuda_generada_bs": float(datos_asistencia['deuda_generada_bs']) if datos_asistencia else 0.0,
                    "modificado_manualmente": bool(datos_asistencia['modificado_manualmente']) if datos_asistencia else False,
                    "observaciones": datos_asistencia['observaciones'] if datos_asistencia else "",
                    
                    # ⚡ Textos Visuales para el HTML
                    "ui_entrada": format_ui_time(datos_asistencia.get('hora_entrada') if datos_asistencia else None),
                    "ui_alm_in": format_ui_time(datos_asistencia.get('hora_inicio_almuerzo') if datos_asistencia else None),
                    "ui_alm_out": format_ui_time(datos_asistencia.get('hora_fin_almuerzo') if datos_asistencia else None),
                    "ui_salida": format_ui_time(datos_asistencia.get('hora_salida') if datos_asistencia else None),
                    
                    # ⚡ Datos Crudos para el Modal de Edición (SweetAlert)
                    "exige_almuerzo_edicion": exige_almuerzo,
                    "raw_entrada": str(datos_asistencia['hora_entrada'])[:5] if datos_asistencia and datos_asistencia.get('hora_entrada') else '',
                    "raw_alm_in": str(datos_asistencia['hora_inicio_almuerzo'])[:5] if datos_asistencia and datos_asistencia.get('hora_inicio_almuerzo') else '',
                    "raw_alm_out": str(datos_asistencia['hora_fin_almuerzo'])[:5] if datos_asistencia and datos_asistencia.get('hora_fin_almuerzo') else '',
                    "raw_salida": str(datos_asistencia['hora_salida'])[:5] if datos_asistencia and datos_asistencia.get('hora_salida') else '',
                }
                if ausencia_activa:
                    datos_modal["ausencia_detalle"] = {
                        "tipo": ausencia_activa["tipo"],
                        "motivo": ausencia_activa.get("motivo", ""),
                        "hora_inicio": str(ausencia_activa["hora_inicio"])[:5] if ausencia_activa.get("hora_inicio") else None,
                        "hora_fin": str(ausencia_activa["hora_fin"])[:5] if ausencia_activa.get("hora_fin") else None,
                        "horas_totales": float(ausencia_activa["horas_totales"] or 0),
                        "requiere_reposicion": bool(ausencia_activa.get("requiere_reposicion"))
                    }

            dias_calendario.append({
                "dia": d, "fecha": fecha_str, "tipo_dia": tipo_dia, "estado_asistencia": estado_asistencia,
                "horas_extras_hoy": h_extras_hoy, "minutos_retraso": int(datos_asistencia['minutos_retraso_entrada']) if datos_asistencia else 0,
                "permite_clic": permite_clic, "tiene_permiso_parcial": bool(ausencia_activa) and tipo_dia == 'laboral',
                "datos_modal": datos_modal
            })

        for a in ausencias_raw:
            if a.get('requiere_reposicion'): sum_horas_reponer += float(a.get('horas_totales') or 0)
                
        saldo_neto = sum_horas_extras - sum_horas_reponer
        estado_bolsillo = "deuda" if saldo_neto < 0 else ("superavit_pagable" if paga_extras and saldo_neto > 0 else ("superavit_no_pagable" if saldo_neto > 0 else "equilibrado"))

        # ⚡ 5. PAQUETE JSON MAESTRO LISTO PARA SER PINTADO ⚡
        return {
            "cabecera": {
                "turno_nombre": emp_data.get('turno_nombre') or 'No asignado',
                "turno_ingreso": str(emp_data['hora_ingreso'])[:5] if emp_data and emp_data.get('hora_ingreso') else '',
                "turno_salida": str(emp_data['hora_salida'])[:5] if emp_data and emp_data.get('hora_salida') else '',
                "turno_almuerzo_texto": f"Incluye {emp_data.get('almuerzo_min')}m Almuerzo" if emp_data and emp_data.get('almuerzo') else "Sin horario de almuerzo"
            },
            "calendario": { "vacios_inicio": vacios_inicio, "dias": dias_calendario },
            "kpis": {
                "dias_trabajados": dias_trabajados_count, "horas_trabajadas": round(sum_horas_trabajadas, 2),
                "horas_esperadas": round(sum_horas_esperadas, 2), "retraso_total_min": retraso_total_min,
                "deuda_total_bs": round(deuda_total_bs, 2)
            },
            "bolsillo": { "saldo_neto": round(saldo_neto, 2), "horas_reponer": round(sum_horas_reponer, 2), "estado": estado_bolsillo }
        }
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 13. MÓDULO: FERIADOS
# ==============================================================================
@app.get("/feriados")
async def obtener_feriados(anio: int = None, usuario = Depends(verificar_token)):
    schema = usuario.get("schema_name")
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        if anio:
            # ⚡ INTELIGENCIA: El motor SQL filtra por año y ordena por mes/día en milisegundos
            query = f"""
                SELECT id, fecha, descripcion, tipo, recurrente 
                FROM {schema}.feriados 
                WHERE eliminado = FALSE 
                AND (recurrente = TRUE OR EXTRACT(YEAR FROM fecha) = %s)
                ORDER BY EXTRACT(MONTH FROM fecha) ASC, EXTRACT(DAY FROM fecha) ASC
            """
            cur.execute(query, (anio,))
        else:
            cur.execute(f"SELECT id, fecha, descripcion, tipo, recurrente FROM {schema}.feriados WHERE eliminado = FALSE ORDER BY fecha ASC")
            
        feriados_db = cur.fetchall()
        
        # Blindaje de formato de fechas para JSON
        for f in feriados_db:
            if f.get("fecha"):
                f.update({"fecha": str(f.get("fecha"))})
        return feriados_db
    finally:
        cur.close()
        conn.close()

@app.post("/sincronizar-feriados/{anio}")
async def sincronizar_feriados_moviles(anio: int, usuario = Depends(verificar_token)):
    schema = usuario.get("schema_name")
    conn = conectar_bd(schema)
    cur = conn.cursor()
    
    try:
        feriados_bolivia = holidays.Bolivia(years=anio)
        insertados = 0

        for fecha, desc in feriados_bolivia.items():
            # Convertimos todo a minúsculas para que no falle por mayúsculas
            desc_lower = desc.lower()
            
            # ⚡ CORRECCIÓN: Buscamos tanto en Inglés como en Español
            es_movil = any(m in desc_lower for m in ["carnival", "carnaval", "good friday", "viernes", "corpus"])
            
            if es_movil:
                # Estandarizamos el nombre para tu base de datos
                desc_es = "Corpus Christi"
                if "carnival" in desc_lower or "carnaval" in desc_lower:
                    desc_es = "Feriado de Carnaval"
                elif "good friday" in desc_lower or "viernes" in desc_lower:
                    desc_es = "Viernes Santo"

                # Verificamos si ya existe antes de insertar
                cur.execute(f"SELECT id FROM {schema}.feriados WHERE fecha = %s AND eliminado = FALSE", (fecha,))
                if not cur.fetchone():
                    cur.execute(f"""
                        INSERT INTO {schema}.feriados (fecha, descripcion, tipo, recurrente)
                        VALUES (%s, %s, 'nacional', FALSE)
                    """, (fecha, desc_es))
                    insertados += 1
        
        conn.commit()
        return {"status": "success", "insertados": insertados}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

class FeriadoCreate(BaseModel):
    fecha: str
    descripcion: str
    tipo: str # Aquí guardaremos 'Nacional', 'La Paz', 'Santa Cruz', etc.
    recurrente: bool = False

@app.post("/feriados")
async def crear_feriado_manual(feriado: FeriadoCreate, usuario = Depends(verificar_token)):
    # Seguridad: Solo admins pueden crear feriados
    if usuario.get("rol") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Sin permisos para crear feriados.")
        
    schema = usuario.get("schema_name")
    conn = conectar_bd(schema)
    cur = conn.cursor()
    
    try:
        # Verificamos que no exista un feriado exactamente en esa fecha
        cur.execute(f"SELECT id FROM {schema}.feriados WHERE fecha = %s AND eliminado = FALSE", (feriado.fecha,))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail="Ya existe un feriado registrado en esta fecha.")

        cur.execute(f"""
            INSERT INTO {schema}.feriados (fecha, descripcion, tipo, recurrente)
            VALUES (%s, %s, %s, %s) RETURNING id
        """, (feriado.fecha, feriado.descripcion, feriado.tipo, feriado.recurrente))
        
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return {"mensaje": "Feriado guardado exitosamente", "id": nuevo_id}
    except psycopg2.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail="Error en la base de datos.")
    finally:
        cur.close()
        conn.close()


@app.delete("/feriados/{feriado_id}")
async def eliminar_feriado(feriado_id: int, usuario = Depends(verificar_token)):
    if usuario.get("rol") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Sin permisos para eliminar feriados.")
        
    schema = usuario.get("schema_name")
    conn = conectar_bd(schema)
    cur = conn.cursor()
    
    try:
        # Seguridad: Evitamos que borren los feriados fijos base (recurrente = TRUE)
        cur.execute(f"SELECT recurrente FROM {schema}.feriados WHERE id = %s", (feriado_id,))
        resultado = cur.fetchone()
        
        if not resultado:
            raise HTTPException(status_code=404, detail="Feriado no encontrado.")
            
        if resultado[0] is True: # Si recurrente es True
            raise HTTPException(status_code=403, detail="Los feriados fijos nacionales no se pueden eliminar.")

        # Borrado Lógico (Soft Delete)
        cur.execute(f"UPDATE {schema}.feriados SET eliminado = TRUE WHERE id = %s", (feriado_id,))
        conn.commit()
        return {"mensaje": "Feriado eliminado exitosamente."}
    except psycopg2.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail="Error en la base de datos.")
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 14. MÓDULO: REPORTE DEL DÍA (GLOBAL) - FILTROS Y ESTADOS INTELIGENTES
# ==============================================================================

@app.get("/reporte-diario/{fecha}")
async def obtener_reporte_diario(
    fecha: str, 
    q: str = "", 
    sucursal_id: str = "", 
    seccion_id: str = "", 
    turno_id: str = "", 
    usuario = Depends(verificar_token)
):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        from datetime import datetime, time
        hoy_srv = datetime.now()
        fecha_req = datetime.strptime(fecha, "%Y-%m-%d").date()
        es_hoy = (fecha_req == hoy_srv.date())
        hora_actual = hoy_srv.time()

        condiciones = ["e.eliminado = FALSE", "e.activo = TRUE"]
        parametros = [fecha, fecha]

        if sucursal_id and sucursal_id.isdigit():
            condiciones.append("e.sucursal_id = %s")
            parametros.append(int(sucursal_id))
        if seccion_id and seccion_id.isdigit():
            condiciones.append("e.seccion_id = %s")
            parametros.append(int(seccion_id))
        if turno_id and turno_id.isdigit():
            condiciones.append("e.turno_id = %s")
            parametros.append(int(turno_id))
        if q:
            condiciones.append("(e.nombres ILIKE %s OR e.apellidos ILIKE %s OR e.ci ILIKE %s)")
            termino = f"%{q}%"
            parametros.extend([termino, termino, termino])

        where_clause = " AND ".join(condiciones)

        cur.execute(f"""
            SELECT 
                e.id, e.nombres, e.apellidos, e.foto_perfil, e.cargo,
                s.nombre as sucursal_nombre, sec.nombre as seccion_nombre,
                t.nombre as turno_nombre, t.hora_ingreso, t.hora_salida, t.almuerzo as turno_almuerzo,
                ad.estado, ad.hora_entrada as marcaje_entrada, ad.hora_salida as marcaje_salida, 
                ad.hora_inicio_almuerzo as marcaje_alm_in, ad.hora_fin_almuerzo as marcaje_alm_out,
                ad.minutos_retraso_entrada, ad.horas_trabajadas,
                (SELECT tipo 
                 FROM {schema}.ausencias a 
                 WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE
                 AND %s BETWEEN a.fecha_inicio AND a.fecha_fin 
                 LIMIT 1) as estado_ausencia
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            LEFT JOIN {schema}.asistencia_diaria ad ON ad.empleado_id = e.id AND ad.fecha = %s
            WHERE {where_clause}
            ORDER BY t.hora_ingreso ASC, e.nombres ASC
        """, tuple(parametros))
        
        reporte = cur.fetchall()

        # ⚡ KPIs SEPARADOS
        kpis = {
            "total": len(reporte),
            "presentes": 0,
            "faltas": 0,
            "vacaciones": 0,
            "permisos": 0,
            "retrasos": 0,
            "turnos_activos": []
        }

        turnos_en_curso = set()

        for fila in reporte:
            h_in = fila.get("hora_ingreso")
            h_out = fila.get("hora_salida")
            
            fila["ui_turno"] = f"{str(h_in)[:5]} a {str(h_out)[:5]}" if h_in and h_out else "Sin Turno"
            fila["ui_entrada"] = str(fila["marcaje_entrada"])[:5] if fila.get("marcaje_entrada") else "--:--"
            fila["ui_salida"] = str(fila["marcaje_salida"])[:5] if fila.get("marcaje_salida") else "--:--"
            
            # ⚡ UI ALMUERZO DINÁMICO
            if fila.get("turno_almuerzo"):
                a_in = str(fila["marcaje_alm_in"])[:5] if fila.get("marcaje_alm_in") else "--:--"
                a_out = str(fila["marcaje_alm_out"])[:5] if fila.get("marcaje_alm_out") else "--:--"
                fila["ui_almuerzo"] = f"{a_in} / {a_out}"
            else:
                fila["ui_almuerzo"] = '<span class="text-gray-400 font-normal">N/A</span>'

            estado_ausencia = fila.get("estado_ausencia")
            estado_ad = fila.get("estado")
            
            ui_estado = "Desconocido"
            ui_color = "bg-gray-100 text-gray-700 border-gray-200"
            ui_icono = "fa-question-circle"

            en_horario_laboral = False
            if es_hoy and h_in and h_out:
                if h_out < h_in: en_horario_laboral = hora_actual >= h_in or hora_actual <= h_out
                else: en_horario_laboral = h_in <= hora_actual <= h_out
                if en_horario_laboral: turnos_en_curso.add(fila["turno_nombre"])

            # ⚡ ÁRBOL DE DECISIONES DE PRIORIDAD TÁCTICA
            if not h_in:
                ui_estado = "Sin Turno"
                ui_color = "bg-slate-100 text-slate-500 border-slate-200"
                ui_icono = "fa-calendar-times"
                
            elif fila.get("marcaje_entrada"):
                # 1. PRIORIDAD: YA ESTÁ FÍSICAMENTE EN LA EMPRESA
                kpis["presentes"] += 1
                if estado_ad in ["Trabajando", "En Curso"]:
                    ui_estado = "En Curso"
                    ui_color = "bg-green-50 text-green-700 border-green-200 shadow-sm"
                    ui_icono = "fa-user-clock"
                elif estado_ad == "Puntual":
                    ui_estado = "Jornada Completada"
                    ui_color = "bg-emerald-50 text-emerald-700 border-emerald-200"
                    ui_icono = "fa-check-circle"
                elif estado_ad == "Tarde":
                    ui_estado = f"Llegó Tarde ({fila['minutos_retraso_entrada']}m)"
                    ui_color = "bg-orange-50 text-orange-700 border-orange-200"
                    ui_icono = "fa-running"
                    kpis["retrasos"] += 1
                else:
                    ui_estado = estado_ad or "Con Registro"
                    ui_color = "bg-gray-50 text-gray-500 border-gray-200"
                    ui_icono = "fa-info-circle"
            else:
                # 2. NO HA MARCADO ENTRADA (Evaluamos Ausencias o Faltas)
                if estado_ausencia == "vacacion":
                    ui_estado = "Vacación"
                    ui_color = "bg-blue-50 text-blue-700 border-blue-200"
                    ui_icono = "fa-umbrella-beach"
                    kpis["vacaciones"] += 1
                elif estado_ausencia == "permiso":
                    # Si tiene permiso pero el motor de cálculo ya lo marcó como "Falta", es que se le pasó la hora del permiso.
                    if estado_ad in ["Falta", "Retraso Crítico"]:
                        ui_estado = "Falta Post-Permiso"
                        ui_color = "bg-red-50 text-red-700 border-red-200"
                        ui_icono = "fa-exclamation-triangle"
                        kpis["faltas"] += 1
                    else:
                        ui_estado = "De Permiso"
                        ui_color = "bg-yellow-50 text-yellow-700 border-yellow-200"
                        ui_icono = "fa-user-md"
                        kpis["permisos"] += 1
                elif es_hoy and hora_actual < h_in:
                    ui_estado = "Próximo a Iniciar"
                    ui_color = "bg-slate-50 text-slate-600 border-slate-200"
                    ui_icono = "fa-clock"
                elif es_hoy and en_horario_laboral:
                    ui_estado = "Retraso Crítico"
                    ui_color = "bg-red-50 text-red-600 border-red-200"
                    ui_icono = "fa-exclamation-triangle"
                    kpis["faltas"] += 1
                elif estado_ad in ["Falta", "Incompleto"]:
                    ui_estado = "Falta Injustificada"
                    ui_color = "bg-red-50 text-red-700 border-red-200"
                    ui_icono = "fa-times-circle"
                    kpis["faltas"] += 1
                else:
                    ui_estado = estado_ad or "Sin Registro"
                    ui_color = "bg-gray-50 text-gray-500 border-gray-200"
                    ui_icono = "fa-minus-circle"

            fila["ui_estado"] = ui_estado
            fila["ui_color"] = ui_color
            fila["ui_icono"] = ui_icono

        kpis["turnos_activos"] = list(turnos_en_curso)

        return {
            "fecha": fecha,
            "es_hoy": es_hoy,
            "kpis": kpis,
            "detalle": reporte
        }
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 15. EDICIÓN MANUAL DE ASISTENCIA (Solo para RRHH / Admin)
# ==============================================================================

@app.put("/asistencia/{empleado_id}/editar-dia")
async def editar_asistencia_manual(empleado_id: int, data: dict, request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    
    if usuario["rol"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Sin permisos para editar el historial.")
        
    admin_password = data.get("admin_password")
    if not admin_password:
        raise HTTPException(status_code=409, detail="Contraseña requerida.")
        
    conn_pub = conectar_bd("public")
    cur_pub = conn_pub.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur_pub.execute("SELECT password_hash FROM usuarios WHERE id = %s", (usuario["id"],))
    user_db = cur_pub.fetchone()
    cur_pub.close()
    conn_pub.close()

    if not user_db or not bcrypt.checkpw(admin_password.encode(), user_db["password_hash"].encode()):
        raise HTTPException(status_code=409, detail="Contraseña incorrecta. Edición denegada.")

    fecha = data.get("fecha")
    h_entrada = data.get("hora_entrada")
    h_salida = data.get("hora_salida")
    h_alm_in = data.get("hora_inicio_almuerzo")
    h_alm_out = data.get("hora_fin_almuerzo")
    justificacion = data.get("justificacion", "Edición manual por RRHH")
    
    if not fecha: raise HTTPException(status_code=400, detail="Fecha requerida.")
    
    from datetime import time, timedelta # ⚡ Aseguramos que timedelta esté disponible
    
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) # Usamos RealDictCursor para traer datos fácilmente
    try:
        device_id = "EDICIÓN-MANUAL"
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
        
        # ⚡ 1. TRAEMOS LA TOPOLOGÍA DEL TURNO PARA SABER SI ES NOCTURNO
        cur.execute(f"""
            SELECT e.bio_id, t.hora_ingreso, t.hora_salida 
            FROM {schema}.empleados e
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.id = %s
        """, (empleado_id,))
        emp_data = cur.fetchone()
        
        bio_id = str(emp_data['bio_id']) if emp_data and emp_data['bio_id'] else "S/N"
        t_ingreso = emp_data['hora_ingreso'] if emp_data else None
        t_salida = emp_data['hora_salida'] if emp_data else None
        
        es_nocturno = False
        if t_ingreso and t_salida and t_salida < t_ingreso:
            es_nocturno = True

        # ⚡ 2. LIMPIEZA ABSOLUTA DE LA VENTANA (Mata marcas erróneas del ZKTeco)
        if es_nocturno:
            inicio_ventana = f"{fecha} 12:00:00"
            fin_ventana = f"{fecha_dt + timedelta(days=1)} 11:59:59"
        else:
            inicio_ventana = f"{fecha} 00:00:00"
            fin_ventana = f"{fecha} 23:59:59"

        # A) Respaldo de auditoría: Guardamos qué había antes de borrar
        cur.execute(f"""
            SELECT fecha_hora, device_no FROM {schema}.eventos_brutos 
            WHERE item = %s AND fecha_hora >= %s AND fecha_hora <= %s
            ORDER BY fecha_hora ASC
        """, (bio_id, inicio_ventana, fin_ventana))
        eventos_viejos = cur.fetchall()
        
        # Armamos un texto: "08:25(Z), 23:37(Z)" (Z=ZKTeco, M=Manual)
        backup_txt = ", ".join([f"{e['fecha_hora'].strftime('%H:%M')}({'M' if e['device_no'] == 'EDICIÓN-MANUAL' else 'Z'})" for e in eventos_viejos])
        observacion_final = justificacion
        if backup_txt:
            observacion_final += f" | Sobreescribió: [{backup_txt}]"

        # B) Borrado total de la ventana para reiniciar el día y matar el error
        cur.execute(f"""
            DELETE FROM {schema}.eventos_brutos 
            WHERE item = %s AND fecha_hora >= %s AND fecha_hora <= %s
        """, (bio_id, inicio_ventana, fin_ventana))

        # ⚡ 3. INYECTOR INTELIGENTE DE FECHAS
        def inyectar_marcaje(hora_str, accion, etiqueta):
            if hora_str:
                h_obj = datetime.strptime(hora_str, "%H:%M").time()
                fecha_asignada = fecha_dt
                
                # Cruce de medianoche en turnos nocturnos
                if es_nocturno and h_obj < time(12, 0):
                    fecha_asignada = fecha_dt + timedelta(days=1)
                    
                fh = f"{fecha_asignada} {hora_str}:00"
                
                cur.execute(f"""
                    INSERT INTO {schema}.eventos_brutos (device_no, item, action, fecha_hora, raw_data)
                    VALUES (%s, %s, %s, %s, %s)
                """, (device_id, bio_id, accion, fh, psycopg2.extras.Json({"raw": etiqueta})))

        # Inyectamos solo la nueva verdad dictada por RRHH
        inyectar_marcaje(h_entrada, '0', "MANUAL-IN")
        inyectar_marcaje(h_alm_in, '1', "MANUAL-LUNCH-OUT")
        inyectar_marcaje(h_alm_out, '0', "MANUAL-LUNCH-IN")
        inyectar_marcaje(h_salida, '1', "MANUAL-OUT")
        
        # Desbloqueamos temporalmente para permitir recálculo
        cur.execute(f"UPDATE {schema}.asistencia_diaria SET modificado_manualmente = FALSE WHERE empleado_id = %s AND fecha = %s", (empleado_id, fecha_dt))
        conn.commit()

        # 4. Despertamos al Cerebro para leer la nueva historia
        exito = procesar_asistencia_dia(schema, empleado_id, fecha_dt)
        
        if exito:
            # ⚡ Sellamos la observación final con el respaldo incluido
            cur.execute(f"""
                UPDATE {schema}.asistencia_diaria 
                SET observaciones = %s, modificado_manualmente = TRUE
                WHERE empleado_id = %s AND fecha = %s
            """, (observacion_final, empleado_id, fecha_dt))
            conn.commit()
            return {"mensaje": "Día corregido. El sistema limpió las marcas erróneas y aplicó la edición."}
        else:
            raise HTTPException(status_code=500, detail="Error en el Motor de Cálculo.")

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/asistencia/{empleado_id}/eliminar-dia/{fecha}")
async def eliminar_asistencia_dia(empleado_id: int, fecha: str, request: Request, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    
    # 1. Verificación de Roles
    if usuario["rol"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Sin permisos para eliminar historial.")
        
    # 2. Captura del Body en la petición DELETE
    try:
        data = await request.json()
    except:
        raise HTTPException(status_code=400, detail="Formato JSON inválido o vacío.")
        
    admin_password = data.get("admin_password")
    if not admin_password:
        raise HTTPException(status_code=409, detail="Contraseña requerida.")
        
    # 3. Verificación de Contraseña en BD Public
    import bcrypt
    conn_pub = conectar_bd("public")
    cur_pub = conn_pub.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur_pub.execute("SELECT password_hash FROM usuarios WHERE id = %s", (usuario["id"],))
    user_db = cur_pub.fetchone()
    cur_pub.close()
    conn_pub.close()

    if not user_db or not bcrypt.checkpw(admin_password.encode(), user_db["password_hash"].encode()):
        raise HTTPException(status_code=409, detail="Contraseña incorrecta. Eliminación denegada.")
    
    # 4. 🛑 Lógica de Eliminación (Solo si pasó los filtros)
    conn = conectar_bd(schema)
    cur = conn.cursor()
    try:
        cur.execute(f"SELECT bio_id FROM {schema}.empleados WHERE id = %s", (empleado_id,))
        emp_data = cur.fetchone()
        bio_id = str(emp_data[0]) if emp_data and emp_data[0] else "S/N"

        # Borramos los eventos brutos
        cur.execute(f"DELETE FROM {schema}.eventos_brutos WHERE item = %s AND DATE(fecha_hora) = %s", (bio_id, fecha))
        
        # Borramos el cálculo diario
        cur.execute(f"DELETE FROM {schema}.asistencia_diaria WHERE empleado_id = %s AND fecha = %s", (empleado_id, fecha))
        
        conn.commit()
        
        # Forzamos al motor a recalcular para que ponga estado "Falta"
        from datetime import datetime
        procesar_asistencia_dia(schema, empleado_id, datetime.strptime(fecha, "%Y-%m-%d").date())
        
        return {"mensaje": "Registros eliminados. El día ha vuelto a estado inicial."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 📊 MOTOR DE REPORTES: EXCEL Y PDF (VERSIÓN DEFINITIVA Y LEGAL)
# ==============================================================================

@app.get("/empleados/{empleado_id}/reporte/excel/{anio}/{mes}")
async def descargar_reporte_excel(empleado_id: int, anio: int, mes: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        import calendar
        from datetime import date
        
        # 1. INFO EMPLEADO Y TURNO
        cur.execute(f"""
            SELECT e.nombres, e.apellidos, e.ci, e.cargo, 
                   s.nombre as sucursal, sec.nombre as seccion, 
                   t.nombre as turno_nombre, t.hora_ingreso, t.hora_salida,
                   t.horas_extras as turno_paga_extras, t.dias as turno_dias
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.id = %s
        """, (empleado_id,))
        emp = cur.fetchone()
        
        paga_extras = emp.get('turno_paga_extras', False)
        raw_dias = emp.get('turno_dias', '{}')
        import json
        dias_laborales = json.loads(raw_dias) if isinstance(raw_dias, str) else (raw_dias or {})

        # 2. ASISTENCIA CACHÉ
        cur.execute(f"""
            SELECT fecha, hora_entrada, hora_salida, horas_trabajadas, horas_extras, 
                   minutos_retraso_entrada, deuda_generada_bs, estado, observaciones, modificado_manualmente
            FROM {schema}.asistencia_diaria
            WHERE empleado_id = %s AND EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s
        """, (empleado_id, anio, mes))
        asist_dict = {str(a['fecha']): a for a in cur.fetchall()}

        # 3. AUSENCIAS
        _, dias_del_mes = calendar.monthrange(anio, mes)
        cur.execute(f"""
            SELECT tipo, fecha_inicio, fecha_fin, motivo, requiere_reposicion, horas_totales
            FROM {schema}.ausencias
            WHERE empleado_id = %s AND estado = 'aprobado' AND eliminado = FALSE
            AND ((EXTRACT(YEAR FROM fecha_inicio) = %s AND EXTRACT(MONTH FROM fecha_inicio) = %s) 
                 OR (EXTRACT(YEAR FROM fecha_fin) = %s AND EXTRACT(MONTH FROM fecha_fin) = %s)
                 OR (fecha_inicio <= %s AND fecha_fin >= %s))
        """, (empleado_id, anio, mes, anio, mes, f"{anio}-{mes:02d}-{dias_del_mes:02d}", f"{anio}-{mes:02d}-01"))
        ausencias = cur.fetchall()

        # 4. FERIADOS
        cur.execute(f"SELECT fecha, descripcion, recurrente FROM {schema}.feriados WHERE eliminado = FALSE")
        feriados_dict = {}
        for f in cur.fetchall():
            fd = str(f['fecha'])
            feriados_dict[fd[5:] if f['recurrente'] else fd] = f['descripcion']

        # 5. VARIABLES DE CÁLCULO
        total_dias = 0; total_horas = 0.0; total_extras = 0.0; total_retraso = 0; total_multa = 0.0
        filas_datos = []
        hoy_srv = date.today()
        mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}

        # ⚡ 6. ITERACIÓN DE DÍA 1 A FIN DE MES
        for d in range(1, dias_del_mes + 1):
            fecha_dt = date(anio, mes, d)
            fecha_str = f"{anio}-{mes:02d}-{d:02d}"
            md_str = f"{mes:02d}-{d:02d}"
            letra_dia = mapa_dias[fecha_dt.weekday()]
            es_laboral = bool(dias_laborales.get(letra_dia, False))
            
            asis = asist_dict.get(fecha_str)
            ausencia = next((a for a in ausencias if str(a['fecha_inicio']) <= fecha_str <= str(a['fecha_fin'])), None)
            feriado = feriados_dict.get(fecha_str) or feriados_dict.get(md_str)

            ent = '--:--'; sal = '--:--'; h_net = 0.0; h_ext = 0.0; ret = 0; mul = 0.0; est = ''; obs = ''

            if asis:
                ent = asis['hora_entrada'].strftime('%H:%M') if asis['hora_entrada'] else '--:--'
                sal = asis['hora_salida'].strftime('%H:%M') if asis['hora_salida'] else '--:--'
                h_net = float(asis['horas_trabajadas'] or 0)
                h_ext = float(asis['horas_extras'] or 0)
                ret = int(asis['minutos_retraso_entrada'] or 0)
                mul = float(asis['deuda_generada_bs'] or 0)
                est = asis['estado']
                obs = asis['observaciones'] or ''
                if asis['modificado_manualmente']: obs = f"[*] {obs}"
                if h_net > 0: total_dias += 1
            else:
                if ausencia:
                    est = 'Vacación' if ausencia['tipo'] == 'vacacion' else 'Permiso'
                    obs = ausencia.get('motivo') or ''
                elif feriado:
                    est = 'Feriado'
                    obs = feriado
                elif not es_laboral:
                    est = 'Descanso'
                elif fecha_dt > hoy_srv:
                    est = '' # Futuro, lo dejamos en blanco
                else:
                    est = 'Falta'

            # Agregamos la advertencia de deuda si existe
            if ausencia and ausencia.get('requiere_reposicion'):
                obs += f" [DEBE REPONER {float(ausencia['horas_totales'])} HRS]"

            total_horas += h_net
            total_extras += h_ext
            total_retraso += ret
            total_multa += mul

            if est: # Solo guardamos si no es un día futuro sin datos
                filas_datos.append([fecha_str, ent, sal, h_net, h_ext, ret, mul, est, obs])

        # 7. DEUDA GLOBAL MENSUAL
        cur.execute(f"""
            SELECT SUM(horas_totales) as deuda_horas FROM {schema}.ausencias
            WHERE empleado_id = %s AND tipo = 'permiso' AND requiere_reposicion = TRUE AND estado = 'aprobado' AND eliminado = FALSE
            AND ((EXTRACT(YEAR FROM fecha_inicio) = %s AND EXTRACT(MONTH FROM fecha_inicio) = %s) 
                 OR (EXTRACT(YEAR FROM fecha_fin) = %s AND EXTRACT(MONTH FROM fecha_fin) = %s))
        """, (empleado_id, anio, mes, anio, mes))
        horas_a_reponer = float(cur.fetchone()['deuda_horas'] or 0)

        # ----------------- ARMADO DEL EXCEL -----------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {mes}-{anio}"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

        # ENCABEZADO
        ws.merge_cells('A1:I1')
        ws['A1'] = f"REPORTE MENSUAL DE ASISTENCIA Y RENDIMIENTO - {mes:02d}/{anio}"
        ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
        ws['A1'].alignment = Alignment(horizontal="center")

        ws['A3'] = "Colaborador:"; ws['B3'] = f"{emp['nombres']} {emp['apellidos']}"
        ws['A4'] = "Doc (C.I.):";  ws['B4'] = emp['ci']
        ws['A5'] = "Cargo:";       ws['B5'] = emp.get('cargo', 'N/A')
        
        ws['D3'] = "Sucursal:";    ws['E3'] = emp.get('sucursal', 'N/A')
        ws['D4'] = "Sección:";     ws['E4'] = emp.get('seccion', 'N/A')
        ws['D5'] = "Turno Base:";  ws['E5'] = f"{emp.get('turno_nombre', 'N/A')}"

        # ⚡ LÓGICA DE EXTRAS VS EXCEDENTE
        tit_extras = "Horas Extras:" if paga_extras else "Tiempo Excedente:"
        col_extras = "Hrs Ext." if paga_extras else "Hrs Exc."
        
        ws['H3'] = "Días Trabajados:"; ws['I3'] = total_dias; ws['I3'].font = Font(bold=True, color="1E3A8A")
        ws['H4'] = "Total Horas:";     ws['I4'] = round(total_horas, 2); ws['I4'].font = Font(bold=True, color="1E3A8A")
        ws['H5'] = tit_extras;         ws['I5'] = round(total_extras, 2); ws['I5'].font = Font(bold=True, color="008000" if paga_extras else "808080")
        ws['H6'] = "Deuda Hrs:";       ws['I6'] = horas_a_reponer; ws['I6'].font = Font(bold=True, color="FF8C00")
        ws['H7'] = "Multas (Bs.):";    ws['I7'] = round(total_multa, 2); ws['I7'].font = Font(bold=True, color="FF0000")

        # CABECERAS DE TABLA
        headers = ["Fecha", "Entrada", "Salida", "Hrs Trab.", col_extras, "Retraso (m)", "Multa (Bs)", "Estado", "Observaciones"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # LLENADO
        fila_actual = 10
        for fila in filas_datos:
            for i, val in enumerate(fila, 1):
                celda = ws.cell(row=fila_actual, column=i, value=val)
                if i < 9: celda.alignment = Alignment(horizontal="center")
            fila_actual += 1

        anchos = [12, 10, 10, 10, 10, 12, 12, 15, 60]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Reporte_Completo_{mes}_{anio}.xlsx"})
    finally:
        cur.close(); conn.close()


@app.get("/empleados/{empleado_id}/reporte/pdf/{anio}/{mes}")
async def descargar_reporte_pdf(empleado_id: int, anio: int, mes: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        import calendar
        from datetime import date
        
        # 1. INFO EMPLEADO
        cur.execute(f"""
            SELECT e.nombres, e.apellidos, e.ci, e.cargo, s.nombre as sucursal, sec.nombre as seccion, 
                   t.nombre as turno, t.hora_ingreso, t.hora_salida, t.horas_extras as turno_paga_extras, t.dias as turno_dias
            FROM {schema}.empleados e
            LEFT JOIN {schema}.sucursales s ON e.sucursal_id = s.id
            LEFT JOIN {schema}.secciones sec ON e.seccion_id = sec.id
            LEFT JOIN {schema}.turnos t ON e.turno_id = t.id
            WHERE e.id = %s
        """, (empleado_id,))
        emp = cur.fetchone()

        paga_extras = emp.get('turno_paga_extras', False)
        raw_dias = emp.get('turno_dias', '{}')
        import json
        dias_laborales = json.loads(raw_dias) if isinstance(raw_dias, str) else (raw_dias or {})

        # 2. DATOS
        cur.execute(f"SELECT * FROM {schema}.asistencia_diaria WHERE empleado_id = %s AND EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s", (empleado_id, anio, mes))
        asist_dict = {str(a['fecha']): a for a in cur.fetchall()}

        _, dias_del_mes = calendar.monthrange(anio, mes)
        cur.execute(f"""
            SELECT tipo, fecha_inicio, fecha_fin, motivo, requiere_reposicion, horas_totales
            FROM {schema}.ausencias
            WHERE empleado_id = %s AND estado = 'aprobado' AND eliminado = FALSE
            AND ((EXTRACT(YEAR FROM fecha_inicio) = %s AND EXTRACT(MONTH FROM fecha_inicio) = %s) 
                 OR (EXTRACT(YEAR FROM fecha_fin) = %s AND EXTRACT(MONTH FROM fecha_fin) = %s)
                 OR (fecha_inicio <= %s AND fecha_fin >= %s))
        """, (empleado_id, anio, mes, anio, mes, f"{anio}-{mes:02d}-{dias_del_mes:02d}", f"{anio}-{mes:02d}-01"))
        ausencias = cur.fetchall()

        cur.execute(f"SELECT fecha, descripcion, recurrente FROM {schema}.feriados WHERE eliminado = FALSE")
        feriados_dict = {}
        for f in cur.fetchall():
            fd = str(f['fecha'])
            feriados_dict[fd[5:] if f['recurrente'] else fd] = f['descripcion']

        # VARIABLES
        t_dias = 0; t_hrs = 0.0; t_ext = 0.0; t_ret = 0; t_mul = 0.0
        datos_tabla = []
        hoy_srv = date.today()
        mapa_dias = {0: 'L', 1: 'M', 2: 'X', 3: 'J', 4: 'V', 5: 'S', 6: 'D'}

        # ⚡ ITERADOR DÍA A DÍA
        for d in range(1, dias_del_mes + 1):
            fecha_dt = date(anio, mes, d)
            fecha_str = f"{anio}-{mes:02d}-{d:02d}"
            md_str = f"{mes:02d}-{d:02d}"
            letra_dia = mapa_dias[fecha_dt.weekday()]
            es_laboral = bool(dias_laborales.get(letra_dia, False))
            
            asis = asist_dict.get(fecha_str)
            ausencia = next((a for a in ausencias if str(a['fecha_inicio']) <= fecha_str <= str(a['fecha_fin'])), None)
            feriado = feriados_dict.get(fecha_str) or feriados_dict.get(md_str)

            ent = '--:--'; sal = '--:--'; h_net = 0.0; h_ext = 0.0; ret = 0; mul = 0.0; est = ''; obs = ''

            if asis:
                ent = asis['hora_entrada'].strftime('%H:%M') if asis['hora_entrada'] else '--:--'
                sal = asis['hora_salida'].strftime('%H:%M') if asis['hora_salida'] else '--:--'
                h_net = float(asis['horas_trabajadas'] or 0)
                h_ext = float(asis['horas_extras'] or 0)
                ret = int(asis['minutos_retraso_entrada'] or 0)
                mul = float(asis['deuda_generada_bs'] or 0)
                est = asis['estado']
                obs = asis['observaciones'] or ''
                if asis['modificado_manualmente']: obs = f"[*] {obs}"
                if h_net > 0: t_dias += 1
            else:
                if ausencia:
                    est = 'Vacación' if ausencia['tipo'] == 'vacacion' else 'Permiso'
                    obs = ausencia.get('motivo') or ''
                elif feriado:
                    est = 'Feriado'
                    obs = feriado
                elif not es_laboral:
                    est = 'Descanso'
                elif fecha_dt > hoy_srv:
                    est = '' 
                else:
                    est = 'Falta'

            # ⚡ ETIQUETAS DE COLOR EN PDF
            if ausencia and ausencia.get('requiere_reposicion'):
                obs += f" <font color='red'><b>[DEBE REPONER {float(ausencia['horas_totales'])} HRS]</b></font>"

            if est == 'Falta': est = f"<font color='red'>{est}</font>"
            elif est == 'Tarde': est = f"<font color='orange'>{est}</font>"
            elif est == 'Vacación': est = f"<font color='#3b82f6'>{est}</font>"
            elif est == 'Permiso': est = f"<font color='#eab308'>{est}</font>"

            multa_str = f"Bs {mul}" if mul > 0 else "-"

            t_hrs += h_net; t_ext += h_ext; t_ret += ret; t_mul += mul

            if est:
                datos_tabla.append([
                    fecha_str, ent, sal, f"{h_net}", f"{h_ext}", f"{ret}m", multa_str,
                    Paragraph(est, getSampleStyleSheet()['Normal']), 
                    Paragraph(obs, getSampleStyleSheet()['Normal'])
                ])

        cur.execute(f"""
            SELECT SUM(horas_totales) as deuda_horas FROM {schema}.ausencias
            WHERE empleado_id = %s AND tipo = 'permiso' AND requiere_reposicion = TRUE AND estado = 'aprobado' AND eliminado = FALSE
            AND ((EXTRACT(YEAR FROM fecha_inicio) = %s AND EXTRACT(MONTH FROM fecha_inicio) = %s) OR (EXTRACT(YEAR FROM fecha_fin) = %s AND EXTRACT(MONTH FROM fecha_fin) = %s))
        """, (empleado_id, anio, mes, anio, mes))
        horas_a_reponer = float(cur.fetchone()['deuda_horas'] or 0)

        # ----------------- ARMADO DEL PDF -----------------
        stream = io.BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        elementos = []
        estilos = getSampleStyleSheet()

        estilo_titulo = estilos['Title']; estilo_titulo.textColor = colors.HexColor("#1E3A8A")
        estilo_normal = estilos['Normal']; estilo_normal.fontSize = 9

        elementos.append(Paragraph(f"<b>REPORTE MENSUAL DE ASISTENCIA Y RENDIMIENTO</b>", estilo_titulo))
        elementos.append(Spacer(1, 10))

        # ⚡ KPI DINÁMICO DE EXTRAS
        if paga_extras:
            kpi_ext_txt = f"<b>Horas Extras:</b> <font color='green'>+{round(t_ext, 2)}</font>"
        else:
            kpi_ext_txt = f"<b>Tiempo Excedente:</b> <font color='gray'>+{round(t_ext, 2)}</font>"

        turno_txt = f"{emp.get('turno', 'N/A')} ({str(emp['hora_ingreso'])[:5] if emp.get('hora_ingreso') else ''} - {str(emp['hora_salida'])[:5] if emp.get('hora_salida') else ''})"
        
        info_data = [
            [Paragraph(f"<b>Colaborador:</b> {emp['nombres']} {emp['apellidos']}", estilo_normal), Paragraph(f"<b>Días Trabajados:</b> {t_dias}", estilo_normal)],
            [Paragraph(f"<b>Documento:</b> {emp['ci']}", estilo_normal), Paragraph(f"<b>Total Horas Netas:</b> {round(t_hrs, 2)} hrs", estilo_normal)],
            [Paragraph(f"<b>Cargo:</b> {emp.get('cargo', 'N/A')} | <b>Sucursal:</b> {emp.get('sucursal', 'N/A')}", estilo_normal), Paragraph(kpi_ext_txt, estilo_normal)],
            [Paragraph(f"<b>Turno Base:</b> {turno_txt}", estilo_normal), Paragraph(f"<b>Horas a Reponer:</b> <font color='orange'>{horas_a_reponer}</font>" if horas_a_reponer > 0 else "<b>Horas a Reponer:</b> 0", estilo_normal)],
            [Paragraph(f"<b>Periodo:</b> {mes:02d}/{anio}", estilo_normal), Paragraph(f"<b>Total Multas:</b> <font color='red'><b>Bs. {round(t_mul, 2)}</b></font> ({t_ret} min retraso)", estilo_normal)]
        ]
        
        t_info = Table(info_data, colWidths=[400, 300])
        t_info.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8FAFC")), ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")), ('PADDING', (0, 0), (-1, -1), 6), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elementos.append(t_info)
        elementos.append(Spacer(1, 15))

        # ⚡ CABECERA MUTANTE DE LA TABLA
        col_extras = "Hrs Ext." if paga_extras else "Hrs Exc."
        datos_matriz = [["Fecha", "Entrada", "Salida", "Hrs Netas", col_extras, "Retraso", "Multa", "Estado", "Observaciones"]] + datos_tabla

        t_datos = Table(datos_matriz, colWidths=[65, 50, 50, 55, 50, 50, 50, 65, 295])
        t_datos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (7, -1), 'CENTER'), ('ALIGN', (8, 0), (8, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, 0), 10), ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]), ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elementos.append(t_datos)
        elementos.append(Spacer(1, 10))
        elementos.append(Paragraph("<font size=7 color=gray>[*] El símbolo asterisco en observaciones indica que los marcajes del día fueron editados o completados manualmente por Recursos Humanos o el Administrador.</font>", estilos['Normal']))
        elementos.append(Spacer(1, 40))
        elementos.append(Table([[Paragraph("<center>_________________________<br/>Firma del Empleado</center>", estilo_normal), Paragraph("<center>_________________________<br/>Autorizado por (RRHH)</center>", estilo_normal)]], colWidths=[365, 365]))

        doc.build(elementos)
        stream.seek(0)
        
        return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=Reporte_Completo_{mes}_{anio}.pdf"})
    finally:
        cur.close(); conn.close()

# ==============================================================================
# 1x. MÓDULO: SIMULADOR DE HARDWARE (DEV / QA)
# ==============================================================================

# ── SIMULADOR DE HARDWARE (VERSIÓN EVENT-DRIVEN) ──
@app.post("/simulador/evento")
async def simulador_evento(data: dict, background_tasks: BackgroundTasks, usuario = Depends(verificar_token)):
    # 1. El Frontend Tonto solo manda 3 textos
    device_no = data.get("device_no")
    bio_id = data.get("bio_id")
    fecha_hora_str = data.get("fecha_hora") # "2023-10-25 08:30:00"

    if not device_no or not bio_id or not fecha_hora_str:
        raise HTTPException(status_code=400, detail="Faltan datos para simular el evento.")

    conn = conectar_bd("public")
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT schema_name FROM public.dispositivos WHERE numero_serie = %s", (device_no,))
        disp = cur.fetchone()
        if not disp:
            raise HTTPException(status_code=404, detail="El lector simulado no existe.")
            
        schema = disp["schema_name"]
        
        # Validamos al empleado en su schema
        cur.execute(f"SELECT id, nombres FROM {schema}.empleados WHERE bio_id = %s", (bio_id,))
        emp = cur.fetchone()
        if not emp:
            raise HTTPException(status_code=404, detail="Ese Bio ID no pertenece a ningún empleado.")

        # 2. Guardar Marcaje Bruto (Caja negra)
        raw_string = f"{bio_id}\t{fecha_hora_str}\t0\t1\tSIMULADO"
        cur.execute(f"""
            INSERT INTO {schema}.eventos_brutos (device_no, item, action, fecha_hora, raw_data)
            VALUES (%s, %s, %s, %s, %s)
        """, (device_no, bio_id, "0", fecha_hora_str, psycopg2.extras.Json({"raw": raw_string})))
        conn.commit()

        # 🚀 3. EL GRAN DISPARADOR EN SEGUNDO PLANO (Velocidad de la luz)
        fecha_dt = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M:%S").date()
        ayer_dt = fecha_dt - timedelta(days=1)

        background_tasks.add_task(procesar_asistencia_dia, schema, emp['id'], fecha_dt)
        background_tasks.add_task(procesar_asistencia_dia, schema, emp['id'], ayer_dt)

        # Respondemos de inmediato al frontend
        return {"mensaje": f"Marcaje inyectado para {emp['nombres']}. El servidor calculará la asistencia en segundo plano."}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 16. MÓDULO: DASHBOARD Y MÉTRICAS TÁCTICAS (FRONTEND TONTO)
# ==============================================================================

@app.get("/dashboard/resumen")
async def obtener_dashboard_resumen(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd(schema)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        from datetime import datetime, date
        hoy_dt = datetime.now()
        hoy = hoy_dt.date()
        ahora = hoy_dt.time()

        # 1. FUERZA LABORAL
        cur.execute(f"SELECT COUNT(*) FILTER (WHERE activo=TRUE) as activos, COUNT(*) FILTER (WHERE activo=FALSE) as inactivos FROM {schema}.empleados WHERE eliminado=FALSE")
        res_personal = cur.fetchone()
        activos = res_personal['activos'] or 0
        inactivos = res_personal['inactivos'] or 0

        # 2. IDENTIFICAR TURNOS EN CURSO (Tiempo Real)
        cur.execute(f"SELECT id, nombre, hora_ingreso, hora_salida FROM {schema}.turnos WHERE eliminado=FALSE")
        todos_turnos = cur.fetchall()
        turnos_activos_ids = []
        turnos_activos_info = []

        for t in todos_turnos:
            h_in = t['hora_ingreso']
            h_out = t['hora_salida']
            is_active = False
            if h_in and h_out:
                if h_out < h_in: # Nocturno
                    is_active = ahora >= h_in or ahora <= h_out
                else: # Diurno
                    is_active = h_in <= ahora <= h_out

            if is_active:
                turnos_activos_ids.append(t['id'])
                # Formatear horas para el frontend
                t['rango'] = f"{str(h_in)[:5]} a {str(h_out)[:5]}"
                turnos_activos_info.append(t)

        # 3. MÉTRICAS TÁCTICAS (Solo de los empleados en el Turno Activo)
        asistencia_stats = {"total_esperado": 0, "en_curso": 0, "tarde": 0, "faltas": 0, "vacaciones": 0, "permisos": 0}

        if turnos_activos_ids:
            ids_tuple = tuple(turnos_activos_ids)
            cur.execute(f"""
                SELECT 
                    e.id, ad.estado, ad.hora_entrada,
                    (SELECT tipo FROM {schema}.ausencias a WHERE a.empleado_id = e.id AND a.estado = 'aprobado' AND a.eliminado = FALSE AND %s BETWEEN a.fecha_inicio AND a.fecha_fin LIMIT 1) as estado_ausencia
                FROM {schema}.empleados e
                LEFT JOIN {schema}.asistencia_diaria ad ON ad.empleado_id = e.id AND ad.fecha = %s
                WHERE e.eliminado = FALSE AND e.activo = TRUE AND e.turno_id IN %s
            """, (hoy, hoy, ids_tuple))

            empleados_activos = cur.fetchall()
            asistencia_stats["total_esperado"] = len(empleados_activos)

            for emp in empleados_activos:
                ausencia = emp['estado_ausencia']
                estado_ad = emp['estado']

                if emp['hora_entrada']:
                    if estado_ad in ["Trabajando", "En Curso", "Puntual"]: asistencia_stats["en_curso"] += 1
                    elif estado_ad == "Tarde": asistencia_stats["tarde"] += 1
                else:
                    if ausencia == "vacacion": asistencia_stats["vacaciones"] += 1
                    elif ausencia == "permiso": asistencia_stats["permisos"] += 1
                    else: asistencia_stats["faltas"] += 1 # Si debía marcar y no lo hizo, o está en Retraso Crítico

        # 4. PRÓXIMOS FERIADOS (Desde hoy en adelante)
        cur.execute(f"SELECT fecha, descripcion, recurrente FROM {schema}.feriados WHERE eliminado = FALSE")
        feriados_db = cur.fetchall()
        proximos_feriados = []

        for f in feriados_db:
            f_date = f['fecha']
            if f['recurrente']:
                try: # Construir fecha para este año
                    f_date_this_year = date(hoy.year, f_date.month, f_date.day)
                    if f_date_this_year >= hoy:
                        proximos_feriados.append({"fecha": f_date_this_year, "descripcion": f['descripcion']})
                    else: # Si ya pasó este año, lo agendamos para el próximo
                        proximos_feriados.append({"fecha": date(hoy.year + 1, f_date.month, f_date.day), "descripcion": f['descripcion']})
                except ValueError: pass
            else:
                if f_date >= hoy:
                    proximos_feriados.append({"fecha": f_date, "descripcion": f['descripcion']})

        # Ordenar cronológicamente y tomar los 4 más próximos
        proximos_feriados.sort(key=lambda x: x['fecha'])
        feriados_formateados = [{"fecha_str": x['fecha'].strftime("%d/%m/%Y"), "descripcion": x['descripcion']} for x in proximos_feriados[:4]]

        return {
            "personal": {"activos": activos, "inactivos": inactivos, "total": activos + inactivos},
            "turnos": {"total_configurados": len(todos_turnos), "activos": turnos_activos_info},
            "asistencia": asistencia_stats,
            "feriados": feriados_formateados
        }
    finally:
        cur.close()
        conn.close()

# ==============================================================================
# 17. MÓDULO: GESTIÓN DE LECTORES BIOMÉTRICOS
# ==============================================================================

@app.get("/lectores")
async def obtener_lectores(usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd("public") # Todo se guarda en la maestra
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Si la tabla no existe, la creamos silenciosamente (Auto-Setup)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.dispositivos (
                id SERIAL PRIMARY KEY,
                numero_serie VARCHAR(100) UNIQUE NOT NULL,
                nombre VARCHAR(100) NOT NULL,
                schema_name VARCHAR(100) NOT NULL,
                activo BOOLEAN DEFAULT TRUE,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

        cur.execute("SELECT id, numero_serie, nombre, activo FROM public.dispositivos WHERE schema_name = %s ORDER BY id ASC", (schema,))
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()

@app.post("/lectores")
async def registrar_lector(data: dict, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    sn = data.get("numero_serie", "").strip()
    nombre = data.get("nombre", "").strip()
    
    if not sn or not nombre:
        raise HTTPException(status_code=400, detail="El Número de Serie (SN) y el Nombre son obligatorios.")

    conn = conectar_bd("public")
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO public.dispositivos (numero_serie, nombre, schema_name)
            VALUES (%s, %s, %s)
        """, (sn, nombre, schema))
        conn.commit()
        return {"mensaje": "Lector registrado correctamente. Listo para recibir marcajes."}
    except psycopg2.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Ese Número de Serie ya está registrado en el sistema global.")
    finally:
        cur.close()
        conn.close()

@app.delete("/lectores/{lector_id}")
async def eliminar_lector(lector_id: int, usuario = Depends(verificar_token)):
    schema = usuario["schema_name"]
    conn = conectar_bd("public")
    cur = conn.cursor()
    try:
        # Solo permite borrarlo si pertenece a su empresa
        cur.execute("DELETE FROM public.dispositivos WHERE id = %s AND schema_name = %s", (lector_id, schema))
        conn.commit()
        return {"mensaje": "Lector eliminado y desvinculado."}
    finally:
        cur.close()
        conn.close()


# ── RUTA DE ESTADO (Para saber si la API está viva) ──
@app.get("/")
def inicio():
    return {"estado": "API funcionando", "version": "2.0 (Multitenant Base)"}